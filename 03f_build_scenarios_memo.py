"""
PayPal (PYPL) - Scenarios + Investment Memo Tabs
=================================================
Final tabs to complete the model:
  1. Scenarios: Bull / Base / Bear with probability-weighted target price
  2. Investment Memo: Executive summary, thesis, risks, recommendation

Run from scripts/ folder AFTER 03e_build_ratios_dcf.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model", "PYPL_Financial_Model.xlsx")

# =============================================================================
# STYLES
# =============================================================================
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="2E4057")
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="2E4057")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_LABEL = Font(name="Arial", size=10, color="000000")
FONT_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
FONT_BLUE = Font(name="Arial", size=10, color="0000FF")
FONT_BLUE_BOLD = Font(name="Arial", size=10, bold=True, color="0000FF")
FONT_GREEN = Font(name="Arial", size=10, color="008000")
FONT_GREEN_BOLD = Font(name="Arial", size=10, bold=True, color="008000")
FONT_RED = Font(name="Arial", size=10, color="CC0000")
FONT_RED_BOLD = Font(name="Arial", size=10, bold=True, color="CC0000")
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="888888")
FONT_MEMO = Font(name="Arial", size=10, color="333333")
FONT_MEMO_BOLD = Font(name="Arial", size=10, bold=True, color="333333")

FILL_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_BULL = PatternFill("solid", fgColor="E8F5E9")
FILL_BASE = PatternFill("solid", fgColor="E3F2FD")
FILL_BEAR = PatternFill("solid", fgColor="FFEBEE")
FILL_INPUT = PatternFill("solid", fgColor="FFFFCC")
FILL_RESULT = PatternFill("solid", fgColor="E8FFE8")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_TARGET = PatternFill("solid", fgColor="FFF3CD")
FILL_MEMO_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_MEMO_SECTION = PatternFill("solid", fgColor="F0F4F8")
FILL_BUY = PatternFill("solid", fgColor="28A745")
FILL_HOLD = PatternFill("solid", fgColor="FFC107")
FILL_SELL = PatternFill("solid", fgColor="DC3545")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=Side(style="medium", color="000000"))
BORDER_THIN = Border(bottom=Side(style="thin", color="CCCCCC"))
BORDER_DOUBLE = Border(bottom=Side(style="double", color="000000"))
BORDER_SECTION = Border(
    top=Side(style="medium", color="2E4057"),
    bottom=Side(style="thin", color="2E4057"),
)
BORDER_ALL = Border(
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)

FMT_PCT = '0.0%'
FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_DOLLAR = '$#,##0.00'
FMT_DOLLAR_INT = '$#,##0'
FMT_MULTIPLE = '0.0"x"'

# =============================================================================
# ROW REFERENCES
# =============================================================================
# DCF tab references
DCF = {
    "wacc": 17,
    "fcf_2026": 22, "fcf_2027": 22, "fcf_2028": 22,  # row, cols C/D/E
    "tgr": 23,
    "net_debt": 35,
    "shares": 38,
    "implied_price": 39,
}

# Income Statement
IS = {"revenue": 7, "op_income": 23, "net_income": 39, "eps": 46, "shares": 45}

# Cash Flow
CF = {"fcf": 49}


# =============================================================================
# BUILD SCENARIOS TAB
# =============================================================================
def build_scenarios(wb):
    if "Scenarios" in wb.sheetnames:
        del wb["Scenarios"]

    dcf_idx = wb.sheetnames.index("DCF")
    ws = wb.create_sheet("Scenarios", dcf_idx + 1)
    ws.sheet_properties.tabColor = "FF6B35"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 5
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 18

    r = 1

    # Title
    ws.merge_cells("B1:E1")
    ws.cell(row=1, column=2, value="PayPal (PYPL) — Scenario Analysis").font = FONT_TITLE
    ws.cell(row=2, column=2, value="Three scenarios with probability-weighted target price").font = FONT_SMALL

    # =========================================================================
    # SCENARIO HEADERS
    # =========================================================================
    r = 4
    headers = [("", None), ("Bull Case", FILL_BULL), ("Base Case", FILL_BASE), ("Bear Case", FILL_BEAR)]
    for i, (label, fill) in enumerate(headers):
        c = ws.cell(row=r, column=2 + i, value=label)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER
        if fill:
            c.fill = fill
            c.font = Font(name="Arial", size=10, bold=True, color="2E4057")

    r = 5

    # =========================================================================
    # SCENARIO NARRATIVES
    # =========================================================================
    ws.row_dimensions[r].height = 45
    ws.cell(row=r, column=2, value="Narrative").font = FONT_BOLD

    c = ws.cell(row=r, column=3, value="New CEO executes turnaround;\nbranded checkout recovers;\nVenmo monetization accelerates")
    c.font = Font(name="Arial", size=9, color="2E7D32"); c.alignment = ALIGN_LEFT_WRAP; c.fill = FILL_BULL

    c = ws.cell(row=r, column=4, value="Modest growth continues;\nbranded checkout stabilizes;\ncost discipline maintained")
    c.font = Font(name="Arial", size=9, color="1565C0"); c.alignment = ALIGN_LEFT_WRAP; c.fill = FILL_BASE

    c = ws.cell(row=r, column=5, value="Branded checkout declines;\nApple/Google take share;\nmargin compression from competition")
    c.font = Font(name="Arial", size=9, color="C62828"); c.alignment = ALIGN_LEFT_WRAP; c.fill = FILL_BEAR

    r += 1

    # =========================================================================
    # SCENARIO ASSUMPTIONS
    # =========================================================================
    r += 1
    ws.merge_cells(f"B{r}:E{r}")
    c = ws.cell(row=r, column=2, value="KEY ASSUMPTIONS (FY2028E)")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1

    # Assumption rows
    assumptions = [
        ("Revenue Growth '26-'28 Avg", "6.0%", "4.0%", "1.5%", FMT_PCT),
        ("FY2028E Revenue ($M)", 38500, 37320, 34800, FMT_CURRENCY),
        ("Operating Margin '28E", "22.0%", "19.5%", "15.5%", FMT_PCT),
        ("FY2028E Net Income ($M)", 6200, 4800, 3200, FMT_CURRENCY),
        ("FY2028E FCF ($M)", 8500, "(from model)", 3500, FMT_CURRENCY),
        ("CapEx ($M/yr avg)", 700, 850, 950, FMT_CURRENCY),
        ("Terminal Growth Rate", "2.5%", "1.5%", "1.0%", FMT_PCT),
        ("WACC", "8.5%", "9.9%", "11.5%", FMT_PCT),
    ]

    assumption_rows = {}
    for label, bull, base, bear, fmt in assumptions:
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).border = BORDER_THIN

        for col_idx, (val, fill) in enumerate([(bull, FILL_BULL), (base, FILL_BASE), (bear, FILL_BEAR)]):
            c = ws.cell(row=r, column=3 + col_idx, value=val)
            c.font = FONT_BLUE_BOLD; c.fill = fill; c.alignment = ALIGN_RIGHT; c.border = BORDER_THIN
            if isinstance(val, (int, float)):
                c.number_format = fmt

        assumption_rows[label] = r
        r += 1

    r += 1

    # =========================================================================
    # VALUATION OUTPUT
    # =========================================================================
    ws.merge_cells(f"B{r}:E{r}")
    c = ws.cell(row=r, column=2, value="VALUATION OUTPUT")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1

    # FCF 2028E for each scenario
    fcf_row = r
    ws.cell(row=r, column=2, value="FY2028E Free Cash Flow ($M)").font = FONT_LABEL
    # Bull: higher NI, lower capex → ~$6.8B FCF
    # Base: from model → link to CF tab
    # Bear: lower NI, higher capex → ~$3.5B FCF
    ws.cell(row=r, column=3, value=8500).font = FONT_BLUE_BOLD
    ws.cell(row=r, column=3).fill = FILL_BULL; ws.cell(row=r, column=3).number_format = FMT_CURRENCY; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT

    c = ws.cell(row=r, column=4, value=f"='Cash Flow'!L{CF['fcf']}")
    c.font = FONT_GREEN_BOLD; c.fill = FILL_BASE; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT

    ws.cell(row=r, column=5, value=3500).font = FONT_BLUE_BOLD
    ws.cell(row=r, column=5).fill = FILL_BEAR; ws.cell(row=r, column=5).number_format = FMT_CURRENCY; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT
    r += 1

    # Terminal Value
    tv_row = r
    ws.cell(row=r, column=2, value="Terminal Value ($M)").font = FONT_LABEL
    # TV = FCF * (1+g) / (WACC - g)
    ws.cell(row=r, column=3, value=f"=C{fcf_row}*(1+0.025)/(0.085-0.025)").font = FONT_BOLD
    ws.cell(row=r, column=3).fill = FILL_BULL; ws.cell(row=r, column=3).number_format = FMT_CURRENCY; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT

    ws.cell(row=r, column=4, value=f"=D{fcf_row}*(1+0.015)/(0.099-0.015)").font = FONT_BOLD
    ws.cell(row=r, column=4).fill = FILL_BASE; ws.cell(row=r, column=4).number_format = FMT_CURRENCY; ws.cell(row=r, column=4).alignment = ALIGN_RIGHT

    ws.cell(row=r, column=5, value=f"=E{fcf_row}*(1+0.01)/(0.115-0.01)").font = FONT_BOLD
    ws.cell(row=r, column=5).fill = FILL_BEAR; ws.cell(row=r, column=5).number_format = FMT_CURRENCY; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT
    r += 1

    # PV of TV (discounted 3 years)
    pvtv_row = r
    ws.cell(row=r, column=2, value="PV of Terminal Value ($M)").font = FONT_LABEL
    ws.cell(row=r, column=3, value=f"=C{tv_row}/(1+0.085)^3").font = FONT_BOLD
    ws.cell(row=r, column=3).fill = FILL_BULL; ws.cell(row=r, column=3).number_format = FMT_CURRENCY; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value=f"=D{tv_row}/(1+0.099)^3").font = FONT_BOLD
    ws.cell(row=r, column=4).fill = FILL_BASE; ws.cell(row=r, column=4).number_format = FMT_CURRENCY; ws.cell(row=r, column=4).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=5, value=f"=E{tv_row}/(1+0.115)^3").font = FONT_BOLD
    ws.cell(row=r, column=5).fill = FILL_BEAR; ws.cell(row=r, column=5).number_format = FMT_CURRENCY; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT
    r += 1

    # PV of projected FCFs (simplified: use 3-year avg FCF discounted)
    pvfcf_row = r
    ws.cell(row=r, column=2, value="PV of Projected FCFs ($M)").font = FONT_LABEL
    # Bull: avg FCF ~$7B ramping to $8.5B
    ws.cell(row=r, column=3, value=f"=6500/(1+0.085)^1+7500/(1+0.085)^2+C{fcf_row}/(1+0.085)^3").font = FONT_BOLD
    ws.cell(row=r, column=3).fill = FILL_BULL; ws.cell(row=r, column=3).number_format = FMT_CURRENCY; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT

    # Base: link to DCF tab projected FCFs
    ws.cell(row=r, column=4, value=f"='Cash Flow'!J{CF['fcf']}/(1+0.099)^1+'Cash Flow'!K{CF['fcf']}/(1+0.099)^2+'Cash Flow'!L{CF['fcf']}/(1+0.099)^3").font = FONT_GREEN_BOLD
    ws.cell(row=r, column=4).fill = FILL_BASE; ws.cell(row=r, column=4).number_format = FMT_CURRENCY; ws.cell(row=r, column=4).alignment = ALIGN_RIGHT

    # Bear: lower FCFs with 11.5% WACC
    ws.cell(row=r, column=5, value=f"=2500/(1+0.115)^1+3000/(1+0.115)^2+E{fcf_row}/(1+0.115)^3").font = FONT_BOLD
    ws.cell(row=r, column=5).fill = FILL_BEAR; ws.cell(row=r, column=5).number_format = FMT_CURRENCY; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT
    r += 1

    # Enterprise Value
    ev_row = r
    ws.cell(row=r, column=2, value="Enterprise Value ($M)").font = FONT_BOLD
    ws.cell(row=r, column=2).border = BORDER_BOTTOM
    for col, fill in [(3, FILL_BULL), (4, FILL_BASE), (5, FILL_BEAR)]:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"={cl}{pvfcf_row}+{cl}{pvtv_row}")
        c.font = FONT_BOLD; c.fill = fill; c.number_format = FMT_CURRENCY
        c.alignment = ALIGN_RIGHT; c.border = BORDER_BOTTOM
    r += 1

    # Net Debt
    nd_row = r
    ws.cell(row=r, column=2, value="(-) Net Debt ($M)").font = FONT_LABEL
    for col, fill in [(3, FILL_BULL), (4, FILL_BASE), (5, FILL_BEAR)]:
        c = ws.cell(row=r, column=col, value=f"=DCF!C{DCF['net_debt']}")
        c.font = FONT_GREEN; c.fill = fill; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT
    r += 1

    # Equity Value
    eqv_row = r
    ws.cell(row=r, column=2, value="Equity Value ($M)").font = FONT_BOLD
    for col, fill in [(3, FILL_BULL), (4, FILL_BASE), (5, FILL_BEAR)]:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"={cl}{ev_row}-{cl}{nd_row}")
        c.font = FONT_BOLD; c.fill = fill; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT
    r += 1

    # Shares
    shares_row = r
    ws.cell(row=r, column=2, value="Diluted Shares (M)").font = FONT_LABEL
    for col, fill in [(3, FILL_BULL), (4, FILL_BASE), (5, FILL_BEAR)]:
        c = ws.cell(row=r, column=col, value=f"=DCF!C{DCF['shares']}")
        c.font = FONT_GREEN; c.fill = fill; c.number_format = '#,##0'; c.alignment = ALIGN_RIGHT
    r += 1

    r += 1

    # =========================================================================
    # IMPLIED SHARE PRICES
    # =========================================================================
    ws.merge_cells(f"B{r}:E{r}")
    c = ws.cell(row=r, column=2, value="IMPLIED SHARE PRICE PER SCENARIO")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1

    # Price per scenario
    price_row = r
    ws.cell(row=r, column=2, value="Implied Share Price").font = Font(name="Arial", size=12, bold=True, color="2E4057")
    for col, fill, font_color in [(3, FILL_BULL, "008000"), (4, FILL_BASE, "1565C0"), (5, FILL_BEAR, "CC0000")]:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"={cl}{eqv_row}/{cl}{shares_row}")
        c.font = Font(name="Arial", size=14, bold=True, color=font_color)
        c.fill = fill; c.number_format = FMT_DOLLAR; c.alignment = ALIGN_CENTER
        c.border = BORDER_DOUBLE
    r += 2

    # =========================================================================
    # PROBABILITY WEIGHTING
    # =========================================================================
    ws.merge_cells(f"B{r}:E{r}")
    c = ws.cell(row=r, column=2, value="PROBABILITY-WEIGHTED TARGET PRICE")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1

    # Probability inputs
    prob_row = r
    ws.cell(row=r, column=2, value="Probability Weight").font = FONT_LABEL

    c = ws.cell(row=r, column=3, value=0.20)
    c.font = FONT_BLUE_BOLD; c.fill = FILL_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_CENTER

    c = ws.cell(row=r, column=4, value=0.40)
    c.font = FONT_BLUE_BOLD; c.fill = FILL_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_CENTER

    c = ws.cell(row=r, column=5, value=0.40)
    c.font = FONT_BLUE_BOLD; c.fill = FILL_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_CENTER
    r += 1

    # Check sum = 100%
    check_row = r
    ws.cell(row=r, column=2, value="  Sum (must = 100%)").font = FONT_SMALL
    c = ws.cell(row=r, column=3, value=f"=C{prob_row}+D{prob_row}+E{prob_row}")
    c.font = FONT_SMALL; c.number_format = FMT_PCT; c.alignment = ALIGN_CENTER
    r += 1

    # Weighted contributions
    contrib_row = r
    ws.cell(row=r, column=2, value="Weighted Contribution").font = FONT_LABEL
    for col, fill in [(3, FILL_BULL), (4, FILL_BASE), (5, FILL_BEAR)]:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"={cl}{price_row}*{cl}{prob_row}")
        c.font = FONT_BOLD; c.fill = fill; c.number_format = FMT_DOLLAR; c.alignment = ALIGN_CENTER
    r += 2

    # TARGET PRICE
    target_row = r
    ws.merge_cells(f"B{r}:B{r+1}")
    ws.cell(row=r, column=2, value="TARGET PRICE").font = Font(name="Arial", size=14, bold=True, color="2E4057")
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"C{r}:E{r}")
    c = ws.cell(row=r, column=3, value=f"=C{contrib_row}+D{contrib_row}+E{contrib_row}")
    c.font = Font(name="Arial", size=22, bold=True, color="2E4057")
    c.number_format = FMT_DOLLAR; c.alignment = ALIGN_CENTER; c.fill = FILL_TARGET
    c.border = BORDER_DOUBLE
    r += 2

    # Current price & upside
    mkt_row = r
    ws.cell(row=r, column=2, value="Current Market Price").font = FONT_BOLD
    c = ws.cell(row=r, column=3, value=40.42)
    c.font = FONT_BLUE_BOLD; c.fill = FILL_INPUT; c.number_format = FMT_DOLLAR; c.alignment = ALIGN_CENTER
    ws.cell(row=r, column=4, value="As of Feb 9, 2026 (editable)").font = FONT_SMALL
    r += 1

    upside_row = r
    ws.cell(row=r, column=2, value="Implied Upside / (Downside)").font = FONT_BOLD
    c = ws.cell(row=r, column=3, value=f"=(C{target_row}/C{mkt_row})-1")
    c.font = Font(name="Arial", size=14, bold=True, color="008000")
    c.number_format = '0.0%'; c.alignment = ALIGN_CENTER; c.fill = FILL_RESULT
    r += 2

    # =========================================================================
    # RECOMMENDATION LOGIC
    # =========================================================================
    ws.merge_cells(f"B{r}:E{r}")
    c = ws.cell(row=r, column=2, value="RECOMMENDATION")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1

    rec_row = r
    ws.cell(row=r, column=2, value="Rating").font = Font(name="Arial", size=12, bold=True, color="2E4057")
    # IF upside > 20% → BUY, IF > -10% → HOLD, ELSE → SELL
    c = ws.cell(row=r, column=3, value=f'=IF(C{upside_row}>0.2,"BUY",IF(C{upside_row}>-0.1,"HOLD","SELL"))')
    c.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    c.alignment = ALIGN_CENTER
    # Can't do conditional formatting with openpyxl easily, so default to green
    c.fill = FILL_BUY
    r += 1

    ws.cell(row=r, column=2, value="Decision Rule:").font = FONT_SMALL
    ws.cell(row=r, column=3, value="Upside > 20% = BUY | > -10% = HOLD | else SELL").font = FONT_SMALL
    r += 2

    # =========================================================================
    # FOOTBALL FIELD SUMMARY
    # =========================================================================
    ws.merge_cells(f"B{r}:E{r}")
    c = ws.cell(row=r, column=2, value="VALUATION RANGE (Football Field)")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1

    ff_data = [
        ("Bear Case DCF", f"=E{price_row}", FILL_BEAR, FONT_RED_BOLD),
        ("Current Market Price", f"=C{mkt_row}", FILL_INPUT, FONT_BLUE_BOLD),
        ("Analyst Consensus (avg)", 62.46, FILL_WHITE, FONT_BOLD),
        ("Probability-Weighted Target", f"=C{target_row}", FILL_TARGET, Font(name="Arial", size=10, bold=True, color="2E4057")),
        ("Base Case DCF", f"=D{price_row}", FILL_BASE, FONT_GREEN_BOLD),
        ("Bull Case DCF", f"=C{price_row}", FILL_BULL, FONT_GREEN_BOLD),
    ]

    ws.cell(row=r, column=2, value="Metric").font = FONT_HEADER; ws.cell(row=r, column=2).fill = FILL_HEADER
    ws.cell(row=r, column=3, value="Price").font = FONT_HEADER; ws.cell(row=r, column=3).fill = FILL_HEADER
    ws.cell(row=r, column=3).alignment = ALIGN_CENTER
    r += 1

    for label, val, fill, font in ff_data:
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL; ws.cell(row=r, column=2).border = BORDER_THIN
        c = ws.cell(row=r, column=3, value=val)
        c.font = font; c.fill = fill; c.number_format = FMT_DOLLAR; c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
        r += 1

    ws.freeze_panes = "B4"
    print("  Scenarios tab built")
    return ws, target_row, upside_row, mkt_row, price_row, prob_row


# =============================================================================
# BUILD INVESTMENT MEMO TAB
# =============================================================================
def build_memo(wb, scen_refs):
    target_row, upside_row, mkt_row, price_row, prob_row = scen_refs

    if "Investment Memo" in wb.sheetnames:
        del wb["Investment Memo"]

    ws = wb.create_sheet("Investment Memo")
    ws.sheet_properties.tabColor = "2E4057"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 85
    ws.column_dimensions["C"].width = 5

    r = 1

    # =========================================================================
    # HEADER
    # =========================================================================
    ws.merge_cells(f"B{r}:B{r}")
    c = ws.cell(row=r, column=2, value="INVESTMENT MEMO — PayPal Holdings, Inc. (NASDAQ: PYPL)")
    c.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    c.fill = FILL_MEMO_HEADER; c.alignment = ALIGN_LEFT
    ws.row_dimensions[r].height = 35
    r += 1

    c = ws.cell(row=r, column=2, value="Equity Research | February 2026 | Analyst: Carlos")
    c.font = Font(name="Arial", size=10, color="FFFFFF"); c.fill = FILL_MEMO_HEADER
    r += 2

    # =========================================================================
    # HELPER
    # =========================================================================
    def section_header(title):
        nonlocal r
        c = ws.cell(row=r, column=2, value=title)
        c.font = Font(name="Arial", size=12, bold=True, color="2E4057")
        c.fill = FILL_MEMO_SECTION
        c.border = Border(bottom=Side(style="medium", color="2E4057"))
        r += 1

    def memo_line(text, bold=False, indent=False):
        nonlocal r
        prefix = "    " if indent else ""
        c = ws.cell(row=r, column=2, value=prefix + text)
        c.font = FONT_MEMO_BOLD if bold else FONT_MEMO
        c.alignment = ALIGN_LEFT_WRAP
        ws.row_dimensions[r].height = max(20, len(text) // 80 * 15 + 20)
        r += 1

    def blank():
        nonlocal r
        r += 1

    # =========================================================================
    # EXECUTIVE SUMMARY
    # =========================================================================
    section_header("EXECUTIVE SUMMARY")

    memo_line("PayPal Holdings, Inc. is the world's largest digital payments platform, processing $1.8 trillion in TPV across ~200 markets in FY2025. "
              "The company is at an inflection point: Q4 2025 results missed expectations, the board replaced CEO Alex Chriss with HP's Enrique Lores, "
              "and 2026 guidance disappointed with flat-to-declining EPS projections. The stock has fallen ~50% from its 52-week high to ~$40.")
    blank()
    memo_line("Despite near-term headwinds, PayPal generates strong free cash flow (~$7B annually), maintains a disciplined $6B/year buyback program, "
              "and owns valuable network effects across 438M active accounts. Venmo revenue grew 20% to $1.7B and BNPL TPV surpassed $40B. "
              "The question is whether the new CEO can stabilize branded checkout and fend off Apple Pay/Google Pay competition.")
    blank()

    # =========================================================================
    # KEY FINANCIALS
    # =========================================================================
    section_header("KEY FINANCIAL HIGHLIGHTS (FY2025 Actual)")
    memo_line("Revenue: $33.2B (+4% YoY) | Operating Income: $6.0B (18.1% margin) | Net Income: $4.3B | EPS: $4.58", bold=True)
    memo_line("Free Cash Flow: $6.6B (19.9% FCF margin) | Cash: $8.0B | LT Debt: $10.0B | Net Debt: $1.9B", bold=True)
    memo_line("Shares outstanding declined from 1,188M (2019) to 941M (2025) via aggressive buybacks — 21% reduction in 6 years", bold=True)
    blank()

    # =========================================================================
    # INVESTMENT THESIS
    # =========================================================================
    section_header("INVESTMENT THESIS")
    memo_line("The bull case rests on four pillars:", bold=True)
    memo_line("1. Valuation gap: Trading at ~8x trailing FCF, PayPal is priced like a declining business while still growing revenue 4%/year and generating world-class cash flow.", indent=True)
    memo_line("2. Buyback machine: At $40/share, the $6B annual buyback retires ~15% of shares per year — a powerful EPS growth driver regardless of top-line performance.", indent=True)
    memo_line("3. Venmo monetization: The 100M-user platform grew revenue 20% and is on track to exceed $2B. Debit card TPV +50%, Pay with Venmo TPV +32%.", indent=True)
    memo_line("4. New leadership catalyst: Enrique Lores brings operational discipline from HP. The board's willingness to change CEOs shows urgency to create value.", indent=True)
    blank()

    # =========================================================================
    # KEY RISKS
    # =========================================================================
    section_header("KEY RISKS")
    memo_line("1. Branded checkout erosion: Online branded checkout growth decelerated to 1% in Q4 from 6% a year earlier. Apple Pay and Google Pay are gaining share at checkout.", indent=True)
    memo_line("2. CEO transition risk: Third CEO in 3 years. Lores has no payments experience. Strategic direction uncertainty during transition could last 12+ months.", indent=True)
    memo_line("3. Macro headwinds: Management noted pressure from lower/middle-income consumers cutting back discretionary spending. Rising competition + weak consumer = margin risk.", indent=True)
    memo_line("4. Take rate compression: Transaction take rate fell 8 basis points to 1.65% in Q4. Strategic repricing and Venmo/enterprise mix shift continue to pressure unit economics.", indent=True)
    memo_line("5. Guidance withdrawn: Management withdrew the 2027 outlook from Investor Day and will only guide one year at a time — a negative signal about visibility.", indent=True)
    blank()

    # =========================================================================
    # VALUATION
    # =========================================================================
    section_header("VALUATION SUMMARY")
    memo_line("Methodology: 3-year DCF (FCFF) with terminal value via Gordon Growth Model. Three scenarios with probability weighting.", bold=True)
    blank()
    memo_line("Bull Case (20% probability): Revenue recovers to 6% growth, margins expand to 22%, FCF ~$8.5B by 2028. WACC 8.5%, TGR 2.5%.", indent=True)
    memo_line("Base Case (40% probability): Revenue grows 3-5%, margins stabilize at ~19.5%, FCF ~$5.3B by 2028. WACC 9.9%, TGR 1.5%.", indent=True)
    memo_line("Bear Case (40% probability): Revenue stagnates at 1-2%, margins compress to 15.5%, FCF ~$3.5B by 2028. WACC 11.5%, TGR 1.0%.", indent=True)
    blank()
    memo_line("Note: The high Bear Case probability (40%) reflects genuine uncertainty around competitive positioning, CEO transition, and consumer spending weakness. "
              "The market appears to be pricing in a scenario close to the Bear Case.", indent=True)
    blank()

    # =========================================================================
    # RECOMMENDATION
    # =========================================================================
    section_header("RECOMMENDATION")
    blank()

    # Dynamic recommendation block
    rec_r = r
    c = ws.cell(row=r, column=2, value=f'=IF(Scenarios!C{upside_row}>0.2,"RECOMMENDATION: BUY — Target Price: "&TEXT(Scenarios!C{target_row},"$#,##0.00")&" | Current: "&TEXT(Scenarios!C{mkt_row},"$#,##0.00")&" | Upside: "&TEXT(Scenarios!C{upside_row},"0.0%"),IF(Scenarios!C{upside_row}>-0.1,"RECOMMENDATION: HOLD — Target Price: "&TEXT(Scenarios!C{target_row},"$#,##0.00")&" | Current: "&TEXT(Scenarios!C{mkt_row},"$#,##0.00")&" | Upside: "&TEXT(Scenarios!C{upside_row},"0.0%"),"RECOMMENDATION: SELL — Target Price: "&TEXT(Scenarios!C{target_row},"$#,##0.00")&" | Current: "&TEXT(Scenarios!C{mkt_row},"$#,##0.00")&" | Downside: "&TEXT(Scenarios!C{upside_row},"0.0%")))')
    c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    c.fill = FILL_BUY  # Default green; actual color depends on formula output
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[r].height = 30
    r += 2

    memo_line("PayPal represents a contrarian value opportunity with significant margin of safety. The stock is priced for permanent decline, "
              "yet the company generates best-in-class FCF, is aggressively reducing shares outstanding, and has multiple levers for revenue diversification "
              "(Venmo, BNPL, enterprise payments, agentic commerce). The primary risk is execution under new leadership.", bold=True)
    blank()
    memo_line("This is NOT a high-conviction BUY given the CEO transition uncertainty. Position sizing should reflect the wide range of outcomes "
              "shown in the scenario analysis. The thesis requires monitoring branded checkout trends and Lores's strategic direction over the next 2-3 quarters.")
    blank()

    # =========================================================================
    # CATALYST TIMELINE
    # =========================================================================
    section_header("CATALYST TIMELINE")
    memo_line("March 1, 2026: Enrique Lores officially starts as CEO", indent=True)
    memo_line("May 2026: Q1 2026 earnings — first report under new leadership direction", indent=True)
    memo_line("H2 2026: Expected strategic review / new Investor Day under Lores", indent=True)
    memo_line("2027: Full year under new CEO — execution proof point", indent=True)
    blank()

    # =========================================================================
    # METHODOLOGY NOTE
    # =========================================================================
    section_header("METHODOLOGY & SOURCES")
    memo_line("Financial data: SEC EDGAR 10-K filings (FY2019-2025), yfinance API, PayPal Q4 2025 earnings release (Feb 3, 2026)")
    memo_line("Model: 3-statement financial model (IS/BS/CF fully linked) with 3-year forecast (FY2026E-2028E)")
    memo_line("Valuation: Discounted Cash Flow (FCFF), Gordon Growth terminal value, WACC via CAPM")
    memo_line("Scenario analysis: Bull/Base/Bear with probability-weighted target price")
    memo_line("Tools: Python (data extraction + SQL loading), Excel (financial model), Power BI (dashboard — forthcoming)")
    blank()

    # Disclaimer
    section_header("DISCLAIMER")
    memo_line("This analysis is prepared for educational and portfolio demonstration purposes only. It does not constitute investment advice. "
              "The author has no position in PYPL and does not intend to initiate one within 72 hours of this publication. "
              "All projections are forward-looking estimates subject to significant uncertainty.")

    ws.freeze_panes = "B3"
    print("  Investment Memo tab built")
    return ws


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"Opening model: {MODEL_PATH}")
    if not os.path.exists(MODEL_PATH):
        print("  File not found!")
        return

    wb = load_workbook(MODEL_PATH)
    print(f"  Loaded. Tabs: {', '.join(wb.sheetnames)}")

    ws_scen, *scen_refs = build_scenarios(wb)
    build_memo(wb, scen_refs)

    wb.save(MODEL_PATH)
    print(f"\n  ✓ Scenarios + Investment Memo tabs added.")
    print(f"  ✓ File: {MODEL_PATH}")
    print(f"  ✓ Final tabs: {', '.join(wb.sheetnames)}")
    print(f"\n  → Open Excel, Ctrl+Shift+F9")
    print(f"  → Review Scenarios tab: check probability-weighted target price")
    print(f"  → Review Investment Memo: professional equity research format")
    print(f"\n  THE MODEL IS COMPLETE.")
    print(f"\n  Remaining project work:")
    print(f"    - Phase 4: Power BI dashboard")
    print(f"    - GitHub documentation")


if __name__ == "__main__":
    main()
