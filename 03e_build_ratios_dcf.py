"""
PayPal (PYPL) - Ratios + DCF Valuation Tabs
============================================
Adds:
  1. Ratios tab — profitability, liquidity, leverage, efficiency, growth
  2. DCF tab — WACC calculation, FCF projection, terminal value, equity value,
     sensitivity table (WACC vs Terminal Growth)

All values are Excel formulas referencing existing tabs.

Run from scripts/ folder AFTER 03d_build_cash_flow.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model", "PYPL_Financial_Model.xlsx")

# =============================================================================
# STYLES
# =============================================================================
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="2E4057")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_YEAR_FC = Font(name="Arial", size=10, bold=True, color="CCDDFF")
FONT_LABEL = Font(name="Arial", size=10, color="000000")
FONT_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
FONT_GREEN = Font(name="Arial", size=10, color="008000")
FONT_GREEN_BOLD = Font(name="Arial", size=10, bold=True, color="008000")
FONT_BLUE = Font(name="Arial", size=10, color="0000FF")
FONT_BLUE_BOLD = Font(name="Arial", size=10, bold=True, color="0000FF")
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="888888")
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="2E4057")
FONT_DCF_LABEL = Font(name="Arial", size=10, color="000000")
FONT_DCF_RESULT = Font(name="Arial", size=12, bold=True, color="2E4057")
FONT_DCF_BIG = Font(name="Arial", size=16, bold=True, color="008000")

FILL_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_ACTUAL = PatternFill("solid", fgColor="F2F2F2")
FILL_FORECAST = PatternFill("solid", fgColor="E8F0FE")
FILL_SUBTOTAL = PatternFill("solid", fgColor="E8E8E8")
FILL_SUBTOTAL_FC = PatternFill("solid", fgColor="D6E4F0")
FILL_INPUT = PatternFill("solid", fgColor="FFFFCC")
FILL_RESULT = PatternFill("solid", fgColor="E8FFE8")
FILL_SENS_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_SENS_CORNER = PatternFill("solid", fgColor="FFC107")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=Side(style="medium", color="000000"))
BORDER_THIN = Border(bottom=Side(style="thin", color="CCCCCC"))
BORDER_DOUBLE = Border(bottom=Side(style="double", color="000000"))
BORDER_SECTION = Border(
    top=Side(style="medium", color="2E4057"),
    bottom=Side(style="thin", color="2E4057"),
)
BORDER_ALL = Border(
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)

FMT_PCT = '0.0%'
FMT_PCT_CALC = '0.0%;(0.0%);"-"'
FMT_MULTIPLE = '0.0"x"'
FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_DOLLAR = '$#,##0.00'
FMT_DOLLAR_INT = '$#,##0'

# =============================================================================
# ROW REFERENCES
# =============================================================================
IS = {"revenue": 7, "cogs": 12, "gross_profit": 15, "opex": 20,
      "op_income": 23, "int_exp": 29, "ebt": 32, "tax": 36, "net_income": 39, "shares": 45, "eps": 46}
BS = {"cash": 7, "tca": 12, "ppe": 16, "ta": 21, "tcl": 29, "ltd": 33, "te": 46, "tle": 48}
CF = {"cfo": 20, "capex": 25, "fcf": 49}
ASSUMP = {"da": 24, "sbc_pct": 15, "revenue": 7, "tax_rate": 18,
           "rf": None, "erp": None, "beta": None, "kd": None, "tax_wacc": None,
           "debt_w": None, "equity_w": None, "tgr": None}

YEARS_ACTUAL = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
YEARS_FORECAST = [2026, 2027, 2028]
ALL_YEARS = YEARS_ACTUAL + YEARS_FORECAST

def year_col(year):
    if year in YEARS_ACTUAL:
        return 3 + YEARS_ACTUAL.index(year)
    return 10 + YEARS_FORECAST.index(year)

def cl(col_num):
    return get_column_letter(col_num)


# =============================================================================
# BUILD RATIOS TAB
# =============================================================================
def build_ratios(wb):
    if "Ratios" in wb.sheetnames:
        del wb["Ratios"]

    cf_idx = wb.sheetnames.index("Cash Flow")
    ws = wb.create_sheet("Ratios", cf_idx + 1)
    ws.sheet_properties.tabColor = "6F42C1"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for c in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[c].width = 13
    ws.column_dimensions["M"].width = 30

    COL_NOTE = 13
    row = [0]
    def sr(r): row[0] = r
    def gr(): return row[0]

    # Title
    sr(1)
    ws.merge_cells("B1:L1")
    ws.cell(row=1, column=2, value="PayPal (PYPL) — Financial Ratio Analysis").font = FONT_TITLE
    ws.cell(row=2, column=2, value="All ratios calculated from linked financial statements").font = FONT_SMALL
    ws.cell(row=2, column=COL_NOTE, value="Formula").font = Font(name="Arial", size=10, bold=True, italic=True, color="888888")

    # Year headers
    ws.cell(row=3, column=2).fill = FILL_HEADER
    ws.cell(row=3, column=COL_NOTE).fill = FILL_HEADER
    for year in YEARS_ACTUAL:
        c = ws.cell(row=3, column=year_col(year), value=str(year))
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER; c.number_format = '@'
    for year in YEARS_FORECAST:
        c = ws.cell(row=3, column=year_col(year), value=f"{year}E")
        c.font = FONT_YEAR_FC; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER
    for year in YEARS_ACTUAL:
        c = ws.cell(row=4, column=year_col(year), value="Actual")
        c.font = Font(name="Arial", size=8, italic=True, color="888888"); c.alignment = ALIGN_CENTER
    for year in YEARS_FORECAST:
        c = ws.cell(row=4, column=year_col(year), value="Forecast")
        c.font = Font(name="Arial", size=8, italic=True, color="0000FF"); c.alignment = ALIGN_CENTER

    sr(5)

    def section(title):
        r = gr() + 1
        ws.merge_cells(f"B{r}:L{r}")
        c = ws.cell(row=r, column=2, value=title)
        c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
        for col in range(3, 14):
            ws.cell(row=r, column=col).fill = FILL_SECTION
            ws.cell(row=r, column=col).border = BORDER_SECTION
        sr(r + 1)

    def ratio_row(label, formula_template, fmt=FMT_PCT_CALC, note="", bold=False):
        """
        formula_template: string with {c} placeholder for column letter, {y} for year
        e.g. "='Income Statement'!{c}{row1}/'Income Statement'!{c}{row2}"
        """
        r = gr()
        ws.cell(row=r, column=2, value=label).font = FONT_BOLD if bold else FONT_LABEL

        for year in ALL_YEARS:
            col = year_col(year)
            c_letter = cl(col)
            formula = formula_template.replace("{c}", c_letter)
            c = ws.cell(row=r, column=col, value=formula)
            if year in YEARS_FORECAST:
                c.font = FONT_GREEN_BOLD if bold else FONT_GREEN
                c.fill = FILL_FORECAST
            else:
                c.font = FONT_BOLD if bold else FONT_LABEL
                c.fill = FILL_ACTUAL
            c.number_format = fmt
            c.alignment = ALIGN_RIGHT

        if note:
            ws.cell(row=r, column=COL_NOTE, value=note).font = FONT_SMALL

        sr(r + 1)
        return r

    def empty():
        sr(gr() + 1)

    # =========================================================================
    # PROFITABILITY RATIOS
    # =========================================================================
    section("PROFITABILITY")

    ratio_row("Gross Margin",
              f"='Income Statement'!{{c}}{IS['gross_profit']}/'Income Statement'!{{c}}{IS['revenue']}",
              note="Gross Profit / Revenue")

    ratio_row("Operating Margin",
              f"='Income Statement'!{{c}}{IS['op_income']}/'Income Statement'!{{c}}{IS['revenue']}",
              note="EBIT / Revenue")

    ratio_row("Net Margin",
              f"='Income Statement'!{{c}}{IS['net_income']}/'Income Statement'!{{c}}{IS['revenue']}",
              note="Net Income / Revenue")

    ratio_row("EBITDA Margin",
              f"=('Income Statement'!{{c}}{IS['op_income']}+Assumptions!{{c}}{ASSUMP['da']})/'Income Statement'!{{c}}{IS['revenue']}",
              note="(EBIT + D&A) / Revenue")

    ratio_row("FCF Margin",
              f"='Cash Flow'!{{c}}{CF['fcf']}/'Income Statement'!{{c}}{IS['revenue']}",
              note="Free Cash Flow / Revenue", bold=True)

    empty()

    # =========================================================================
    # RETURN RATIOS
    # =========================================================================
    section("RETURNS")

    ratio_row("Return on Equity (ROE)",
              f"='Income Statement'!{{c}}{IS['net_income']}/'Balance Sheet'!{{c}}{BS['te']}",
              note="Net Income / Total Equity", bold=True)

    ratio_row("Return on Assets (ROA)",
              f"='Income Statement'!{{c}}{IS['net_income']}/'Balance Sheet'!{{c}}{BS['ta']}",
              note="Net Income / Total Assets")

    ratio_row("ROIC",
              f"='Income Statement'!{{c}}{IS['op_income']}*(1-Assumptions!{{c}}{ASSUMP['tax_rate']})/('Balance Sheet'!{{c}}{BS['ta']}-'Balance Sheet'!{{c}}{BS['cash']})",
              note="NOPAT / Invested Capital", bold=True)

    empty()

    # =========================================================================
    # LIQUIDITY
    # =========================================================================
    section("LIQUIDITY")

    ratio_row("Current Ratio",
              f"='Balance Sheet'!{{c}}{BS['tca']}/'Balance Sheet'!{{c}}{BS['tcl']}",
              fmt=FMT_MULTIPLE, note="Current Assets / Current Liabilities")

    empty()

    # =========================================================================
    # LEVERAGE
    # =========================================================================
    section("LEVERAGE")

    ratio_row("Debt / Equity",
              f"='Balance Sheet'!{{c}}{BS['ltd']}/'Balance Sheet'!{{c}}{BS['te']}",
              fmt=FMT_MULTIPLE, note="LT Debt / Total Equity")

    ratio_row("Net Debt / EBITDA",
              f"=('Balance Sheet'!{{c}}{BS['ltd']}-'Balance Sheet'!{{c}}{BS['cash']})/('Income Statement'!{{c}}{IS['op_income']}+Assumptions!{{c}}{ASSUMP['da']})",
              fmt=FMT_MULTIPLE, note="(LT Debt - Cash) / EBITDA")

    ratio_row("Interest Coverage",
              f"='Income Statement'!{{c}}{IS['op_income']}/ABS('Income Statement'!{{c}}{IS['int_exp']})",
              fmt=FMT_MULTIPLE, note="EBIT / Interest Expense")

    empty()

    # =========================================================================
    # EFFICIENCY
    # =========================================================================
    section("EFFICIENCY")

    ratio_row("Asset Turnover",
              f"='Income Statement'!{{c}}{IS['revenue']}/'Balance Sheet'!{{c}}{BS['ta']}",
              fmt=FMT_MULTIPLE, note="Revenue / Total Assets")

    ratio_row("CapEx / Revenue",
              f"=ABS('Cash Flow'!{{c}}{CF['capex']})/'Income Statement'!{{c}}{IS['revenue']}",
              note="Capital intensity")

    empty()

    # =========================================================================
    # GROWTH
    # =========================================================================
    section("GROWTH")

    # Revenue growth (skip 2019 — no prior year in model)
    r = gr()
    ws.cell(row=r, column=2, value="Revenue Growth (YoY)").font = FONT_BOLD
    for i, year in enumerate(ALL_YEARS):
        if i == 0:
            continue
        col = year_col(year)
        prev_col = year_col(ALL_YEARS[i-1])
        formula = f"=('Income Statement'!{cl(col)}{IS['revenue']}-'Income Statement'!{cl(prev_col)}{IS['revenue']})/'Income Statement'!{cl(prev_col)}{IS['revenue']}"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FONT_GREEN_BOLD if year in YEARS_FORECAST else FONT_BOLD
        c.fill = FILL_FORECAST if year in YEARS_FORECAST else FILL_ACTUAL
        c.number_format = FMT_PCT_CALC; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=COL_NOTE, value="YoY Revenue Growth").font = FONT_SMALL
    sr(r + 1)

    # Net Income growth
    r = gr()
    ws.cell(row=r, column=2, value="Net Income Growth (YoY)").font = FONT_LABEL
    for i, year in enumerate(ALL_YEARS):
        if i == 0:
            continue
        col = year_col(year)
        prev_col = year_col(ALL_YEARS[i-1])
        formula = f"=('Income Statement'!{cl(col)}{IS['net_income']}-'Income Statement'!{cl(prev_col)}{IS['net_income']})/ABS('Income Statement'!{cl(prev_col)}{IS['net_income']})"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FONT_GREEN if year in YEARS_FORECAST else FONT_LABEL
        c.fill = FILL_FORECAST if year in YEARS_FORECAST else FILL_ACTUAL
        c.number_format = FMT_PCT_CALC; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=COL_NOTE, value="YoY Net Income Growth").font = FONT_SMALL
    sr(r + 1)

    # EPS growth
    r = gr()
    ws.cell(row=r, column=2, value="EPS Growth (YoY)").font = FONT_BOLD
    for i, year in enumerate(ALL_YEARS):
        if i == 0:
            continue
        col = year_col(year)
        prev_col = year_col(ALL_YEARS[i-1])
        formula = f"=('Income Statement'!{cl(col)}{IS['eps']}-'Income Statement'!{cl(prev_col)}{IS['eps']})/ABS('Income Statement'!{cl(prev_col)}{IS['eps']})"
        c = ws.cell(row=r, column=col, value=formula)
        c.font = FONT_GREEN_BOLD if year in YEARS_FORECAST else FONT_BOLD
        c.fill = FILL_FORECAST if year in YEARS_FORECAST else FILL_ACTUAL
        c.number_format = FMT_PCT_CALC; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=COL_NOTE, value="YoY EPS Growth (key for investors)").font = FONT_SMALL
    sr(r + 1)

    ws.freeze_panes = "C5"
    print("  ✓ Ratios tab built")
    return ws


# =============================================================================
# BUILD DCF TAB
# =============================================================================
def build_dcf(wb):
    if "DCF" in wb.sheetnames:
        del wb["DCF"]

    ratios_idx = wb.sheetnames.index("Ratios")
    ws = wb.create_sheet("DCF", ratios_idx + 1)
    ws.sheet_properties.tabColor = "DC3545"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 30

    r = 1

    # =========================================================================
    # SECTION 1: WACC CALCULATION
    # =========================================================================
    ws.merge_cells("B1:E1")
    ws.cell(row=1, column=2, value="PayPal (PYPL) — DCF Valuation").font = FONT_TITLE
    ws.cell(row=2, column=2, value="Discounted Cash Flow — Free Cash Flow to Firm (FCFF)").font = FONT_SMALL

    r = 4
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="WACC CALCULATION")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION

    r = 5
    # WACC inputs — blue font, yellow bg (editable)
    wacc_inputs = [
        ("Risk-Free Rate (Rf)", 0.042, FMT_PCT, "10Y UST yield"),
        ("Equity Risk Premium (ERP)", 0.055, FMT_PCT, "Damodaran 2025"),
        ("Beta (β)", 1.40, "0.00", "Elevated: CEO change + competition"),
        ("Cost of Equity (Ke)", None, FMT_PCT, "CAPM formula"),
        ("", None, None, ""),
        ("Pre-Tax Cost of Debt (Kd)", 0.042, FMT_PCT, "Int Exp / Avg Debt"),
        ("Tax Rate (t)", 0.175, FMT_PCT, "Effective tax rate"),
        ("After-Tax Cost of Debt", None, FMT_PCT, "Kd × (1 - t)"),
        ("", None, None, ""),
        ("Debt Weight (D/V)", 0.30, FMT_PCT, "Target capital structure"),
        ("Equity Weight (E/V)", 0.70, FMT_PCT, "1 - Debt Weight"),
        ("", None, None, ""),
        ("WACC", None, FMT_PCT, "THE discount rate"),
    ]

    input_rows = {}
    for label, val, fmt, note in wacc_inputs:
        if label == "":
            r += 1
            continue

        ws.cell(row=r, column=2, value=label).font = FONT_LABEL

        if val is not None:
            c = ws.cell(row=r, column=3, value=val)
            c.font = FONT_BLUE_BOLD
            c.fill = FILL_INPUT
            c.number_format = fmt
            c.alignment = ALIGN_RIGHT
        
        ws.cell(row=r, column=4, value=note).font = FONT_SMALL

        input_rows[label] = r
        r += 1

    # FORMULAS for calculated fields
    rf_r = input_rows["Risk-Free Rate (Rf)"]
    erp_r = input_rows["Equity Risk Premium (ERP)"]
    beta_r = input_rows["Beta (β)"]
    ke_r = input_rows["Cost of Equity (Ke)"]
    kd_r = input_rows["Pre-Tax Cost of Debt (Kd)"]
    tax_r = input_rows["Tax Rate (t)"]
    atkd_r = input_rows["After-Tax Cost of Debt"]
    dw_r = input_rows["Debt Weight (D/V)"]
    ew_r = input_rows["Equity Weight (E/V)"]
    wacc_r = input_rows["WACC"]

    # Cost of Equity = Rf + Beta × ERP
    c = ws.cell(row=ke_r, column=3, value=f"=C{rf_r}+C{beta_r}*C{erp_r}")
    c.font = FONT_BOLD; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT; c.fill = FILL_RESULT

    # After-Tax Cost of Debt = Kd × (1 - t)
    c = ws.cell(row=atkd_r, column=3, value=f"=C{kd_r}*(1-C{tax_r})")
    c.font = FONT_BOLD; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT; c.fill = FILL_RESULT

    # WACC = E/V × Ke + D/V × Kd × (1-t)
    c = ws.cell(row=wacc_r, column=3, value=f"=C{ew_r}*C{ke_r}+C{dw_r}*C{kd_r}*(1-C{tax_r})")
    c.font = Font(name="Arial", size=12, bold=True, color="DC3545")
    c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT; c.fill = FILL_RESULT

    r += 1

    # =========================================================================
    # SECTION 2: FREE CASH FLOW PROJECTION
    # =========================================================================
    r += 1
    ws.merge_cells(f"B{r}:F{r}")
    c = ws.cell(row=r, column=2, value="FREE CASH FLOW PROJECTION")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 7):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION

    r += 1
    fcf_header_row = r

    # Headers: Year columns for 2026E, 2027E, 2028E, Terminal
    headers = ["", "2026E", "2027E", "2028E", "Terminal"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER
    r += 1

    # FCF row — link to Cash Flow tab
    fcf_proj_row = r
    ws.cell(row=r, column=2, value="Free Cash Flow ($M)").font = FONT_BOLD
    ws.cell(row=r, column=3, value=f"='Cash Flow'!J{CF['fcf']}").font = FONT_GREEN_BOLD
    ws.cell(row=r, column=3).number_format = FMT_CURRENCY; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value=f"='Cash Flow'!K{CF['fcf']}").font = FONT_GREEN_BOLD
    ws.cell(row=r, column=4).number_format = FMT_CURRENCY; ws.cell(row=r, column=4).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=5, value=f"='Cash Flow'!L{CF['fcf']}").font = FONT_GREEN_BOLD
    ws.cell(row=r, column=5).number_format = FMT_CURRENCY; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT

    r += 1

    # Terminal Growth Rate
    tgr_row = r
    ws.cell(row=r, column=2, value="Terminal Growth Rate").font = FONT_LABEL
    c = ws.cell(row=r, column=3, value=0.015)
    c.font = FONT_BLUE_BOLD; c.fill = FILL_INPUT; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value="Key assumption — sensitivity tested below").font = FONT_SMALL
    r += 1

    # Terminal Value = FCF_terminal × (1+g) / (WACC - g)
    tv_row = r
    ws.cell(row=r, column=2, value="Terminal Value ($M)").font = FONT_BOLD
    c = ws.cell(row=r, column=6, value=f"=E{fcf_proj_row}*(1+C{tgr_row})/(C{wacc_r}-C{tgr_row})")
    c.font = FONT_GREEN_BOLD; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_RESULT
    ws.cell(row=r, column=7, value="FCF_2028 × (1+g) / (WACC-g)").font = FONT_SMALL
    r += 1

    # Discount factors
    r += 1
    df_row = r
    ws.cell(row=r, column=2, value="Discount Factor").font = FONT_LABEL
    ws.cell(row=r, column=3, value=f"=1/(1+C{wacc_r})^1").font = FONT_LABEL
    ws.cell(row=r, column=3).number_format = "0.0000"; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value=f"=1/(1+C{wacc_r})^2").font = FONT_LABEL
    ws.cell(row=r, column=4).number_format = "0.0000"; ws.cell(row=r, column=4).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=5, value=f"=1/(1+C{wacc_r})^3").font = FONT_LABEL
    ws.cell(row=r, column=5).number_format = "0.0000"; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=6, value=f"=1/(1+C{wacc_r})^3").font = FONT_LABEL
    ws.cell(row=r, column=6).number_format = "0.0000"; ws.cell(row=r, column=6).alignment = ALIGN_RIGHT
    r += 1

    # PV of FCFs
    pv_row = r
    ws.cell(row=r, column=2, value="PV of Free Cash Flow ($M)").font = FONT_BOLD
    ws.cell(row=r, column=3, value=f"=C{fcf_proj_row}*C{df_row}").font = FONT_GREEN
    ws.cell(row=r, column=3).number_format = FMT_CURRENCY; ws.cell(row=r, column=3).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value=f"=D{fcf_proj_row}*D{df_row}").font = FONT_GREEN
    ws.cell(row=r, column=4).number_format = FMT_CURRENCY; ws.cell(row=r, column=4).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=5, value=f"=E{fcf_proj_row}*E{df_row}").font = FONT_GREEN
    ws.cell(row=r, column=5).number_format = FMT_CURRENCY; ws.cell(row=r, column=5).alignment = ALIGN_RIGHT
    # PV of Terminal Value
    ws.cell(row=r, column=6, value=f"=F{tv_row}*F{df_row}").font = FONT_GREEN_BOLD
    ws.cell(row=r, column=6).number_format = FMT_CURRENCY; ws.cell(row=r, column=6).alignment = ALIGN_RIGHT
    ws.cell(row=r, column=6).fill = FILL_RESULT

    r += 2

    # =========================================================================
    # SECTION 3: EQUITY VALUE BRIDGE
    # =========================================================================
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="EQUITY VALUE BRIDGE")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 6):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION

    r += 1
    # Sum of PV of FCFs
    sum_pv_row = r
    ws.cell(row=r, column=2, value="PV of Projected FCFs ($M)").font = FONT_LABEL
    c = ws.cell(row=r, column=3, value=f"=SUM(C{pv_row}:E{pv_row})")
    c.font = FONT_BOLD; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT
    r += 1

    pv_tv_row = r
    ws.cell(row=r, column=2, value="PV of Terminal Value ($M)").font = FONT_LABEL
    c = ws.cell(row=r, column=3, value=f"=F{pv_row}")
    c.font = FONT_BOLD; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT
    r += 1

    ev_row = r
    ws.cell(row=r, column=2, value="Enterprise Value ($M)").font = FONT_BOLD
    c = ws.cell(row=r, column=3, value=f"=C{sum_pv_row}+C{pv_tv_row}")
    c.font = Font(name="Arial", size=11, bold=True, color="2E4057")
    c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_RESULT
    c.border = BORDER_BOTTOM
    r += 1

    # TV as % of EV
    ws.cell(row=r, column=2, value="  Terminal Value as % of EV").font = FONT_SMALL
    c = ws.cell(row=r, column=3, value=f"=C{pv_tv_row}/C{ev_row}")
    c.font = FONT_SMALL; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value="Should be 60-85% for mature company").font = FONT_SMALL
    r += 2

    # EV to Equity bridge
    net_debt_row = r
    ws.cell(row=r, column=2, value="(−) Net Debt ($M)").font = FONT_LABEL
    c = ws.cell(row=r, column=3, value=f"='Balance Sheet'!I{BS['ltd']}-'Balance Sheet'!I{BS['cash']}")
    c.font = FONT_GREEN; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value="LT Debt - Cash (latest actual: FY2025)").font = FONT_SMALL
    r += 1

    equity_val_row = r
    ws.cell(row=r, column=2, value="Equity Value ($M)").font = Font(name="Arial", size=11, bold=True, color="2E4057")
    c = ws.cell(row=r, column=3, value=f"=C{ev_row}-C{net_debt_row}")
    c.font = Font(name="Arial", size=12, bold=True, color="008000")
    c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_RESULT
    c.border = BORDER_DOUBLE
    r += 2

    # Shares outstanding
    shares_row = r
    ws.cell(row=r, column=2, value="Diluted Shares Outstanding (M)").font = FONT_LABEL
    c = ws.cell(row=r, column=3, value=f"='Income Statement'!I{IS['shares']}")
    c.font = FONT_GREEN; c.number_format = '#,##0'; c.alignment = ALIGN_RIGHT
    ws.cell(row=r, column=4, value="Latest actual (FY2025)").font = FONT_SMALL
    r += 1

    # IMPLIED SHARE PRICE
    price_row = r
    ws.cell(row=r, column=2, value="Implied Share Price").font = Font(name="Arial", size=14, bold=True, color="2E4057")
    c = ws.cell(row=r, column=3, value=f"=C{equity_val_row}/C{shares_row}")
    c.font = Font(name="Arial", size=18, bold=True, color="008000")
    c.number_format = FMT_DOLLAR; c.alignment = ALIGN_RIGHT
    c.fill = FILL_RESULT; c.border = BORDER_DOUBLE
    r += 2

    # =========================================================================
    # SECTION 4: SENSITIVITY TABLE (WACC vs Terminal Growth)
    # =========================================================================
    r += 1
    ws.merge_cells(f"B{r}:H{r}")
    c = ws.cell(row=r, column=2, value="SENSITIVITY ANALYSIS — Implied Share Price")
    c.font = FONT_SECTION; c.fill = FILL_SECTION; c.border = BORDER_SECTION
    for col in range(3, 9):
        ws.cell(row=r, column=col).fill = FILL_SECTION; ws.cell(row=r, column=col).border = BORDER_SECTION
    r += 1
    ws.cell(row=r, column=2, value="WACC (rows) vs Terminal Growth Rate (columns)").font = FONT_SMALL
    r += 1

    # Sensitivity parameters
    wacc_values = [-0.020, -0.010, -0.005, 0, 0.005, 0.010, 0.020]  # offsets from base
    tgr_values = [-0.010, -0.005, 0, 0.005, 0.010]  # offsets from base

    sens_start_row = r
    # Corner cell
    c = ws.cell(row=r, column=2, value="WACC \\ TGR →")
    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF"); c.fill = FILL_SENS_CORNER; c.alignment = ALIGN_CENTER
    c.border = BORDER_ALL

    # TGR column headers (C through G)
    for j, tgr_off in enumerate(tgr_values):
        col = 3 + j
        c = ws.cell(row=r, column=col, value=f"=C{tgr_row}+{tgr_off}")
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = FILL_SENS_HEADER; c.number_format = FMT_PCT; c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL

    r += 1

    # Data rows
    for i, wacc_off in enumerate(wacc_values):
        # WACC row header
        c = ws.cell(row=r, column=2, value=f"=C{wacc_r}+{wacc_off}")
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = FILL_SENS_HEADER; c.number_format = FMT_PCT; c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL

        for j, tgr_off in enumerate(tgr_values):
            col = 3 + j
            # Recalculate: TV = FCF_2028*(1+tgr)/(wacc-tgr), then discount
            # Full formula inline for each cell
            wacc_cell = f"$B{r}"  # This row's WACC
            tgr_cell = f"{cl(col)}${sens_start_row}"  # This col's TGR

            # TV = FCF_2028 * (1+tgr) / (wacc - tgr)
            # PV_TV = TV / (1+wacc)^3
            # PV_FCF = sum of FCF/(1+wacc)^n for n=1,2,3
            # EV = PV_FCFs + PV_TV
            # Equity = EV - Net Debt
            # Price = Equity / Shares

            formula = (
                f"=("
                f"C{fcf_proj_row}/(1+{wacc_cell})^1"
                f"+D{fcf_proj_row}/(1+{wacc_cell})^2"
                f"+E{fcf_proj_row}/(1+{wacc_cell})^3"
                f"+E{fcf_proj_row}*(1+{tgr_cell})/({wacc_cell}-{tgr_cell})/(1+{wacc_cell})^3"
                f"-C{net_debt_row}"
                f")/C{shares_row}"
            )

            c = ws.cell(row=r, column=col, value=formula)
            c.number_format = FMT_DOLLAR
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL

            # Highlight the base case (wacc_off=0, tgr_off=0)
            if wacc_off == 0 and tgr_off == 0:
                c.fill = FILL_SENS_CORNER
                c.font = Font(name="Arial", size=10, bold=True, color="000000")
            else:
                c.font = Font(name="Arial", size=10, color="000000")

        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Yellow cell = base case (matches implied share price above)").font = FONT_SMALL
    r += 1
    ws.cell(row=r, column=2, value="This table shows how sensitive the valuation is to WACC and terminal growth assumptions").font = FONT_SMALL

    ws.freeze_panes = "B4"

    print(f"  ✓ DCF tab built")
    print(f"    WACC row: {wacc_r}")
    print(f"    FCF projection row: {fcf_proj_row}")
    print(f"    Terminal Value row: {tv_row}")
    print(f"    Implied Price row: {price_row}")
    print(f"    Sensitivity table starts: row {sens_start_row}")

    return ws


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"Opening model: {MODEL_PATH}")
    if not os.path.exists(MODEL_PATH):
        print("  File not found!")
        return

    wb = load_workbook(MODEL_PATH)
    print(f"  Loaded. Tabs: {', '.join(wb.sheetnames)}")

    build_ratios(wb)
    build_dcf(wb)

    wb.save(MODEL_PATH)
    print(f"\n  ✓ Ratios + DCF tabs added.")
    print(f"  ✓ File: {MODEL_PATH}")
    print(f"\n  → Open Excel, Ctrl+Shift+F9")
    print(f"  → Check DCF implied share price")
    print(f"  → Review sensitivity table")
    print(f"\n  Next: Scenarios tab + Investment Memo")


if __name__ == "__main__":
    main()
