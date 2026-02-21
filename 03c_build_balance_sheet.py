"""
PayPal (PYPL) - Balance Sheet Tab Builder
==========================================
Opens existing model and adds the Balance Sheet tab.

Key linkages:
  - PP&E forecast: Prior PP&E + CapEx - D&A (Assumptions tab)
  - Retained Earnings: Prior RE + Net Income (IS tab) - Buybacks (Assumptions)
  - Long-Term Debt: from Assumptions tab
  - Cash: placeholder for now → will link to Cash Flow statement
  - Balance Check row: must be $0

Run from scripts/ folder AFTER 03b_build_income_statement.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model", "PYPL_Financial_Model.xlsx")

# =============================================================================
# STYLES
# =============================================================================
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="2E4057")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_YEAR_FC = Font(name="Arial", size=10, bold=True, color="CCDDFF")
FONT_LABEL = Font(name="Arial", size=10, color="000000")
FONT_LABEL_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
FONT_FORMULA = Font(name="Arial", size=10, color="000000")
FONT_GREEN = Font(name="Arial", size=10, color="008000")
FONT_GREEN_BOLD = Font(name="Arial", size=10, bold=True, color="008000")
FONT_BLUE = Font(name="Arial", size=10, color="0000FF")
FONT_BLUE_BOLD = Font(name="Arial", size=10, bold=True, color="0000FF")
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="888888")
FONT_CHECK_OK = Font(name="Arial", size=10, bold=True, color="008000")
FONT_CHECK_FAIL = Font(name="Arial", size=10, bold=True, color="FF0000")

FILL_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_ACTUAL = PatternFill("solid", fgColor="F2F2F2")
FILL_FORECAST = PatternFill("solid", fgColor="E8F0FE")
FILL_SUBTOTAL = PatternFill("solid", fgColor="E8E8E8")
FILL_SUBTOTAL_FC = PatternFill("solid", fgColor="D6E4F0")
FILL_INPUT = PatternFill("solid", fgColor="FFFFCC")
FILL_CHECK = PatternFill("solid", fgColor="E8FFE8")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=Side(style="medium", color="000000"))
BORDER_THIN = Border(bottom=Side(style="thin", color="CCCCCC"))
BORDER_DOUBLE = Border(bottom=Side(style="double", color="000000"))
BORDER_SECTION = Border(
    top=Side(style="medium", color="2E4057"),
    bottom=Side(style="thin", color="2E4057"),
)

FMT_CURRENCY = '#,##0;(#,##0);"-"'

# =============================================================================
# ROW REFERENCES TO OTHER TABS
# =============================================================================
# Income Statement row references (from 03b script output)
IS_ROWS = {
    "revenue": 7,
    "net_income": 39,
    "shares": 45,
}

# Assumptions tab row references
ASSUMP_ROWS = {
    "revenue": 7,
    "capex": 22,
    "da_absolute": 24,
    "long_term_debt": 31,
    "buybacks": 34,
}

# =============================================================================
# HISTORICAL DATA (USD Millions)
# =============================================================================
YEARS_ACTUAL = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
YEARS_FORECAST = [2026, 2027, 2028]

BS = {
    # CURRENT ASSETS
    "Cash & Cash Equivalents": {
        2019: 7349, 2020: 4794, 2021: 5197, 2022: 7776,
        2023: 9081, 2024: 6662, 2025: 8049,
    },
    "Short-Term Investments": {
        2019: 3412, 2020: 8289, 2021: 109, 2022: 3075,
        2023: 4976, 2024: 4261, 2025: 2373,
    },
    "Accounts Receivable": {
        2019: 435, 2020: 577, 2021: 12723, 2022: 2201,
        2023: 1882, 2024: 1581, 2025: 1909,
    },
    "Customer Accounts & Other Current": {
        2019: 27299, 2020: 37335, 2021: 0, 2022: 44372,
        2023: 46630, 2024: 45702, 2025: 47428,
    },
    "Total Current Assets": {
        2019: 38495, 2020: 50995, 2021: 18029, 2022: 57424,
        2023: 62569, 2024: 58206, 2025: 59759,
    },
    # NON-CURRENT ASSETS
    "Property & Equipment, Net": {
        2019: 1693, 2020: 1807, 2021: 1909, 2022: 2201,
        2023: 1882, 2024: 1581, 2025: 1909,
    },
    "Goodwill": {
        2019: 6212, 2020: 9135, 2021: 11454, 2022: 11209,
        2023: 11026, 2024: 10837, 2025: 10864,
    },
    "Intangible Assets": {
        2019: 778, 2020: 1048, 2021: 1332, 2022: 1730,
        2023: 1488, 2024: 1508, 2025: 1700,
    },
    "Other Non-Current Assets": {
        2019: 4155, 2020: 7394, 2021: 43079, 2022: 6060,
        2023: 5201, 2024: 6593, 2025: 5941,
    },
    "Total Assets": {
        2019: 51333, 2020: 70379, 2021: 75803, 2022: 78624,
        2023: 82166, 2024: 78725, 2025: 80173,
    },
    # CURRENT LIABILITIES
    "Accounts Payable": {
        2019: 232, 2020: 252, 2021: 197, 2022: 40140,
        2023: 42074, 2024: 39898, 2025: 40438,
    },
    "Accrued Expenses & Other Current": {
        2019: 26687, 2020: 38195, 2021: 42832, 2022: 4868,
        2023: 6392, 2024: 5592, 2025: 6005,
    },
    "Total Current Liabilities": {
        2019: 26919, 2020: 38447, 2021: 43029, 2022: 45008,
        2023: 48466, 2024: 45490, 2025: 46443,
    },
    # NON-CURRENT LIABILITIES
    "Long-Term Debt": {
        2019: 4965, 2020: 8939, 2021: 8049, 2022: 10417,
        2023: 9676, 2024: 9879, 2025: 9987,
    },
    "Other Non-Current Liabilities": {
        2019: 2520, 2020: 2930, 2021: 2998, 2022: 2925,
        2023: 2973, 2024: 2939, 2025: 3487,
    },
    "Total Liabilities": {
        2019: 34404, 2020: 50316, 2021: 54076, 2022: 58350,
        2023: 61115, 2024: 58308, 2025: 59917,
    },
    # EQUITY
    "Common Stock & APIC": {
        2019: 15588, 2020: 16644, 2021: 0, 2022: 18327,
        2023: 19642, 2024: 20705, 2025: 21582,
    },
    "Retained Earnings": {
        2019: 8342, 2020: 12366, 2021: 16535, 2022: 18954,
        2023: 23200, 2024: 27347, 2025: 32470,
    },
    "Treasury Stock": {
        2019: -6872, 2020: -8507, 2021: 0, 2022: -16079,
        2023: -21045, 2024: -27085, 2025: -33138,
    },
    "AOCI & Other": {
        2019: -173, 2020: -484, 2021: -808, 2022: -928,
        2023: -746, 2024: -550, 2025: -658,
    },
    "Total Stockholders' Equity": {
        2019: 16885, 2020: 20019, 2021: 21727, 2022: 20274,
        2023: 21051, 2024: 20417, 2025: 20256,
    },
    "Total Liabilities & Equity": {
        2019: 51333, 2020: 70379, 2021: 75803, 2022: 78624,
        2023: 82166, 2024: 78725, 2025: 80173,
    },
}


def year_col(year):
    if year in YEARS_ACTUAL:
        return 3 + YEARS_ACTUAL.index(year)
    return 10 + YEARS_FORECAST.index(year)


def col_letter(col_num):
    return chr(64 + col_num)


# =============================================================================
# BUILD BALANCE SHEET
# =============================================================================
def build_balance_sheet(wb):
    if "Balance Sheet" in wb.sheetnames:
        del wb["Balance Sheet"]

    # Insert after Income Statement
    is_idx = wb.sheetnames.index("Income Statement")
    ws = wb.create_sheet("Balance Sheet", is_idx + 1)
    ws.sheet_properties.tabColor = "17A2B8"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for cl in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[cl].width = 14
    ws.column_dimensions["M"].width = 35

    COL_NOTE = 13
    current_row = [0]  # Use list for mutability in nested functions

    def set_row(r):
        current_row[0] = r

    def get_row():
        return current_row[0]

    # --- Title ---
    set_row(1)
    ws.merge_cells("B1:L1")
    ws.cell(row=1, column=2, value="PayPal (PYPL) — Consolidated Balance Sheet").font = FONT_TITLE
    ws.cell(row=2, column=2, value="USD Millions ($M) — As of December 31").font = FONT_SMALL
    ws.cell(row=2, column=COL_NOTE, value="Notes").font = Font(name="Arial", size=10, bold=True, italic=True, color="888888")

    # --- Year Headers (Row 3) ---
    ws.cell(row=3, column=2).fill = FILL_HEADER
    ws.cell(row=3, column=COL_NOTE).fill = FILL_HEADER
    for year in YEARS_ACTUAL:
        c = ws.cell(row=3, column=year_col(year), value=str(year))
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER; c.number_format = '@'
    for year in YEARS_FORECAST:
        c = ws.cell(row=3, column=year_col(year), value=f"{year}E")
        c.font = FONT_YEAR_FC; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER

    # Period labels
    for year in YEARS_ACTUAL:
        c = ws.cell(row=4, column=year_col(year), value="Actual")
        c.font = Font(name="Arial", size=8, italic=True, color="888888"); c.alignment = ALIGN_CENTER
    for year in YEARS_FORECAST:
        c = ws.cell(row=4, column=year_col(year), value="Forecast")
        c.font = Font(name="Arial", size=8, italic=True, color="0000FF"); c.alignment = ALIGN_CENTER

    set_row(5)

    # =========================================================================
    # HELPERS
    # =========================================================================
    def section(title):
        r = get_row() + 1
        ws.merge_cells(f"B{r}:L{r}")
        c = ws.cell(row=r, column=2, value=title)
        c.font = Font(name="Arial", size=11, bold=True, color="2E4057")
        c.fill = FILL_SECTION; c.border = BORDER_SECTION
        for col in range(3, 14):
            ws.cell(row=r, column=col).fill = FILL_SECTION
            ws.cell(row=r, column=col).border = BORDER_SECTION
        set_row(r + 1)

    def row_data(label, data_key=None, forecast=None, note="",
                 subtotal=False, total=False, forecast_type="formula"):
        """
        forecast_type:
          'formula' → green font (cross-sheet formula)
          'input' → blue font, yellow bg (manual input)
        """
        r = get_row()

        # Label
        c = ws.cell(row=r, column=2, value=label)
        c.font = FONT_LABEL_BOLD if (subtotal or total) else FONT_LABEL
        c.alignment = ALIGN_LEFT

        border = BORDER_DOUBLE if total else (BORDER_BOTTOM if subtotal else BORDER_THIN)

        # Actuals
        if data_key and data_key in BS:
            for year in YEARS_ACTUAL:
                val = BS[data_key].get(year)
                if val is not None:
                    c = ws.cell(row=r, column=year_col(year), value=val)
                    c.font = Font(name="Arial", size=10, bold=subtotal or total, color="000000")
                    c.number_format = FMT_CURRENCY
                    c.alignment = ALIGN_RIGHT
                    c.fill = FILL_SUBTOTAL if (subtotal or total) else FILL_ACTUAL
                    c.border = border

        # Forecast
        if forecast:
            for year in YEARS_FORECAST:
                f = forecast.get(year)
                if f is not None:
                    c = ws.cell(row=r, column=year_col(year), value=f)
                    if forecast_type == "input":
                        c.font = FONT_BLUE_BOLD if (subtotal or total) else FONT_BLUE
                        c.fill = FILL_INPUT
                    else:
                        c.font = FONT_GREEN_BOLD if (subtotal or total) else FONT_GREEN
                        c.fill = FILL_SUBTOTAL_FC if (subtotal or total) else FILL_FORECAST
                    c.number_format = FMT_CURRENCY
                    c.alignment = ALIGN_RIGHT
                    c.border = border

        if note:
            ws.cell(row=r, column=COL_NOTE, value=note).font = FONT_SMALL

        set_row(r + 1)
        return r

    def empty():
        set_row(get_row() + 1)

    # =========================================================================
    # ASSETS
    # =========================================================================
    section("CURRENT ASSETS")

    # Cash — will be linked to CF statement later
    cash_row = row_data(
        "Cash & Cash Equivalents",
        data_key="Cash & Cash Equivalents",
        forecast={2026: 8500, 2027: 9200, 2028: 10000},
        forecast_type="input",
        note="PLACEHOLDER — will link to Cash Flow statement",
    )

    st_inv_row = row_data(
        "Short-Term Investments",
        data_key="Short-Term Investments",
        forecast={
            2026: f"=I{0}",
            2027: f"=J{0}",
            2028: f"=K{0}",
        },
        note="Held constant (no material change assumed)",
    )
    ws.cell(row=st_inv_row, column=year_col(2026)).value = f"=I{st_inv_row}"
    ws.cell(row=st_inv_row, column=year_col(2027)).value = f"=J{st_inv_row}"
    ws.cell(row=st_inv_row, column=year_col(2028)).value = f"=K{st_inv_row}"

    # AR forecast: grow at ~revenue growth rate
    ar_row = row_data(
        "Accounts Receivable",
        data_key="Accounts Receivable",
        forecast={
            2026: f"=I{0}*(1+Assumptions!J{ASSUMP_ROWS['revenue']}/Assumptions!I{ASSUMP_ROWS['revenue']}-1)",
            2027: f"=J{0}*(1+Assumptions!K{ASSUMP_ROWS['revenue']}/Assumptions!J{ASSUMP_ROWS['revenue']}-1)",
            2028: f"=K{0}*(1+Assumptions!L{ASSUMP_ROWS['revenue']}/Assumptions!K{ASSUMP_ROWS['revenue']}-1)",
        },
        note="Grows proportionally with revenue",
    )
    # Fix AR formulas with actual row reference
    for year in YEARS_FORECAST:
        col = year_col(year)
        old_formula = ws.cell(row=ar_row, column=col).value
        if old_formula:
            prev_col_l = col_letter(col - 1) if year == 2026 else col_letter(col - 1)
            ws.cell(row=ar_row, column=col).value = old_formula.replace(
                f"I{0}", f"I{ar_row}" if year == 2026 else ""
            )

    # Simpler AR formulas
    ws.cell(row=ar_row, column=year_col(2026)).value = f"=I{ar_row}*('Income Statement'!J{IS_ROWS['revenue']}/'Income Statement'!I{IS_ROWS['revenue']})"
    ws.cell(row=ar_row, column=year_col(2027)).value = f"=J{ar_row}*('Income Statement'!K{IS_ROWS['revenue']}/'Income Statement'!J{IS_ROWS['revenue']})"
    ws.cell(row=ar_row, column=year_col(2028)).value = f"=K{ar_row}*('Income Statement'!L{IS_ROWS['revenue']}/'Income Statement'!K{IS_ROWS['revenue']})"

    cust_row = row_data(
        "Customer Accounts & Other Current",
        data_key="Customer Accounts & Other Current",
        forecast={
            2026: f"=I{0}*1.03",
            2027: f"=J{0}*1.03",
            2028: f"=K{0}*1.03",
        },
        note="Grows ~3% (tied to payment volume growth)",
    )
    # Fix formulas
    ws.cell(row=cust_row, column=year_col(2026)).value = f"=I{cust_row}*1.03"
    ws.cell(row=cust_row, column=year_col(2027)).value = f"=J{cust_row}*1.03"
    ws.cell(row=cust_row, column=year_col(2028)).value = f"=K{cust_row}*1.03"

    empty()

    # Total Current Assets = sum of above
    tca_row = row_data(
        "Total Current Assets",
        data_key="Total Current Assets",
        forecast={
            2026: f"=SUM(J{cash_row}:J{cust_row})",
            2027: f"=SUM(K{cash_row}:K{cust_row})",
            2028: f"=SUM(L{cash_row}:L{cust_row})",
        },
        subtotal=True,
        note="Sum of current asset items",
    )

    empty()

    # --- NON-CURRENT ASSETS ---
    section("NON-CURRENT ASSETS")

    # PP&E: Prior + CapEx - D&A
    ppe_row = row_data(
        "Property & Equipment, Net",
        data_key="Property & Equipment, Net",
        forecast={
            2026: f"=I{0}+Assumptions!J{ASSUMP_ROWS['capex']}-Assumptions!J{ASSUMP_ROWS['da_absolute']}",
            2027: f"=J{0}+Assumptions!K{ASSUMP_ROWS['capex']}-Assumptions!K{ASSUMP_ROWS['da_absolute']}",
            2028: f"=K{0}+Assumptions!L{ASSUMP_ROWS['capex']}-Assumptions!L{ASSUMP_ROWS['da_absolute']}",
        },
        note="Prior PP&E + CapEx - D&A",
    )
    # Fix self-references
    ws.cell(row=ppe_row, column=year_col(2026)).value = f"=I{ppe_row}+Assumptions!J{ASSUMP_ROWS['capex']}-Assumptions!J{ASSUMP_ROWS['da_absolute']}"
    ws.cell(row=ppe_row, column=year_col(2027)).value = f"=J{ppe_row}+Assumptions!K{ASSUMP_ROWS['capex']}-Assumptions!K{ASSUMP_ROWS['da_absolute']}"
    ws.cell(row=ppe_row, column=year_col(2028)).value = f"=K{ppe_row}+Assumptions!L{ASSUMP_ROWS['capex']}-Assumptions!L{ASSUMP_ROWS['da_absolute']}"

    # Goodwill: constant
    gw_row = row_data(
        "Goodwill",
        data_key="Goodwill",
        forecast={
            2026: f"=I{0}",
            2027: f"=J{0}",
            2028: f"=K{0}",
        },
        note="Held constant (no acquisitions assumed)",
    )
    ws.cell(row=gw_row, column=year_col(2026)).value = f"=I{gw_row}"
    ws.cell(row=gw_row, column=year_col(2027)).value = f"=J{gw_row}"
    ws.cell(row=gw_row, column=year_col(2028)).value = f"=K{gw_row}"

    # Intangibles: declining (amortization)
    intang_row = row_data(
        "Intangible Assets",
        data_key="Intangible Assets",
        forecast={
            2026: f"=I{0}",
            2027: f"=J{0}",
            2028: f"=K{0}",
        },
        note="Held constant (no acquisitions assumed)",
    )
    ws.cell(row=intang_row, column=year_col(2026)).value = f"=I{intang_row}"
    ws.cell(row=intang_row, column=year_col(2027)).value = f"=J{intang_row}"
    ws.cell(row=intang_row, column=year_col(2028)).value = f"=K{intang_row}"

    # Other non-current
    other_nca_row = row_data(
        "Other Non-Current Assets",
        data_key="Other Non-Current Assets",
        forecast={
            2026: f"=I{0}",
            2027: f"=J{0}",
            2028: f"=K{0}",
        },
        note="Held stable; immaterial changes assumed",
    )
    ws.cell(row=other_nca_row, column=year_col(2026)).value = f"=I{other_nca_row}"
    ws.cell(row=other_nca_row, column=year_col(2027)).value = f"=J{other_nca_row}"
    ws.cell(row=other_nca_row, column=year_col(2028)).value = f"=K{other_nca_row}"

    empty()

    # Total Assets
    ta_row = row_data(
        "Total Assets",
        data_key="Total Assets",
        forecast={
            2026: f"=J{tca_row}+J{ppe_row}+J{gw_row}+J{intang_row}+J{other_nca_row}",
            2027: f"=K{tca_row}+K{ppe_row}+K{gw_row}+K{intang_row}+K{other_nca_row}",
            2028: f"=L{tca_row}+L{ppe_row}+L{gw_row}+L{intang_row}+L{other_nca_row}",
        },
        total=True,
        note="Current Assets + Non-Current Assets",
    )

    empty()
    empty()

    # =========================================================================
    # LIABILITIES
    # =========================================================================
    section("CURRENT LIABILITIES")

    # AP: grows with revenue
    ap_row = row_data(
        "Accounts Payable",
        data_key="Accounts Payable",
        forecast={
            2026: f"=I{0}*('Income Statement'!J{IS_ROWS['revenue']}/'Income Statement'!I{IS_ROWS['revenue']})",
            2027: f"=J{0}*('Income Statement'!K{IS_ROWS['revenue']}/'Income Statement'!J{IS_ROWS['revenue']})",
            2028: f"=K{0}*('Income Statement'!L{IS_ROWS['revenue']}/'Income Statement'!K{IS_ROWS['revenue']})",
        },
        note="Grows proportionally with revenue",
    )
    ws.cell(row=ap_row, column=year_col(2026)).value = f"=I{ap_row}*('Income Statement'!J{IS_ROWS['revenue']}/'Income Statement'!I{IS_ROWS['revenue']})"
    ws.cell(row=ap_row, column=year_col(2027)).value = f"=J{ap_row}*('Income Statement'!K{IS_ROWS['revenue']}/'Income Statement'!J{IS_ROWS['revenue']})"
    ws.cell(row=ap_row, column=year_col(2028)).value = f"=K{ap_row}*('Income Statement'!L{IS_ROWS['revenue']}/'Income Statement'!K{IS_ROWS['revenue']})"

    # Accrued & Other Current
    acc_row = row_data(
        "Accrued Expenses & Other Current",
        data_key="Accrued Expenses & Other Current",
        forecast={
            2026: f"=I{0}*1.03",
            2027: f"=J{0}*1.03",
            2028: f"=K{0}*1.03",
        },
        note="Grows ~3% annually; estimate",
    )
    ws.cell(row=acc_row, column=year_col(2026)).value = f"=I{acc_row}*1.03"
    ws.cell(row=acc_row, column=year_col(2027)).value = f"=J{acc_row}*1.03"
    ws.cell(row=acc_row, column=year_col(2028)).value = f"=K{acc_row}*1.03"

    empty()

    tcl_row = row_data(
        "Total Current Liabilities",
        data_key="Total Current Liabilities",
        forecast={
            2026: f"=J{ap_row}+J{acc_row}",
            2027: f"=K{ap_row}+K{acc_row}",
            2028: f"=L{ap_row}+L{acc_row}",
        },
        subtotal=True,
        note="Sum of current liability items",
    )

    empty()

    # --- NON-CURRENT LIABILITIES ---
    section("NON-CURRENT LIABILITIES")

    ltd_row = row_data(
        "Long-Term Debt",
        data_key="Long-Term Debt",
        forecast={
            2026: f"=Assumptions!J{ASSUMP_ROWS['long_term_debt']}",
            2027: f"=Assumptions!K{ASSUMP_ROWS['long_term_debt']}",
            2028: f"=Assumptions!L{ASSUMP_ROWS['long_term_debt']}",
        },
        note="From Assumptions tab (gradual deleveraging)",
    )

    other_ncl_row = row_data(
        "Other Non-Current Liabilities",
        data_key="Other Non-Current Liabilities",
        forecast={
            2026: f"=I{0}",
            2027: f"=J{0}",
            2028: f"=K{0}",
        },
        note="Held stable",
    )
    ws.cell(row=other_ncl_row, column=year_col(2026)).value = f"=I{other_ncl_row}"
    ws.cell(row=other_ncl_row, column=year_col(2027)).value = f"=J{other_ncl_row}"
    ws.cell(row=other_ncl_row, column=year_col(2028)).value = f"=K{other_ncl_row}"

    empty()

    tl_row = row_data(
        "Total Liabilities",
        data_key="Total Liabilities",
        forecast={
            2026: f"=J{tcl_row}+J{ltd_row}+J{other_ncl_row}",
            2027: f"=K{tcl_row}+K{ltd_row}+K{other_ncl_row}",
            2028: f"=L{tcl_row}+L{ltd_row}+L{other_ncl_row}",
        },
        subtotal=True,
        note="Current Liab + Non-Current Liab",
    )

    empty()
    empty()

    # =========================================================================
    # EQUITY
    # =========================================================================
    section("STOCKHOLDERS' EQUITY")

    # Common Stock & APIC: grows by SBC
    csapic_row = row_data(
        "Common Stock & APIC",
        data_key="Common Stock & APIC",
        forecast={
            2026: f"=I{0}+Assumptions!J{ASSUMP_ROWS['revenue']}*Assumptions!J15",
            2027: f"=J{0}+Assumptions!K{ASSUMP_ROWS['revenue']}*Assumptions!K15",
            2028: f"=K{0}+Assumptions!L{ASSUMP_ROWS['revenue']}*Assumptions!L15",
        },
        note="Prior + SBC (from Assumptions SBC pct x Revenue)",
    )
    # Fix: SBC % is on row 15 of Assumptions
    ws.cell(row=csapic_row, column=year_col(2026)).value = f"=I{csapic_row}+Assumptions!J{ASSUMP_ROWS['revenue']}*Assumptions!J15"
    ws.cell(row=csapic_row, column=year_col(2027)).value = f"=J{csapic_row}+Assumptions!K{ASSUMP_ROWS['revenue']}*Assumptions!K15"
    ws.cell(row=csapic_row, column=year_col(2028)).value = f"=K{csapic_row}+Assumptions!L{ASSUMP_ROWS['revenue']}*Assumptions!L15"

    # Retained Earnings: Prior + Net Income (buybacks flow through Treasury Stock only)
    re_row = row_data(
        "Retained Earnings",
        data_key="Retained Earnings",
        forecast={
            2026: f"=I{0}+'Income Statement'!J{IS_ROWS['net_income']}",
            2027: f"=J{0}+'Income Statement'!K{IS_ROWS['net_income']}",
            2028: f"=K{0}+'Income Statement'!L{IS_ROWS['net_income']}",
        },
        note="Prior RE + Net Income (buybacks via Treasury Stock)",
    )
    ws.cell(row=re_row, column=year_col(2026)).value = f"=I{re_row}+'Income Statement'!J{IS_ROWS['net_income']}"
    ws.cell(row=re_row, column=year_col(2027)).value = f"=J{re_row}+'Income Statement'!K{IS_ROWS['net_income']}"
    ws.cell(row=re_row, column=year_col(2028)).value = f"=K{re_row}+'Income Statement'!L{IS_ROWS['net_income']}"

    # Treasury Stock: grows by buybacks
    ts_row = row_data(
        "Treasury Stock",
        data_key="Treasury Stock",
        forecast={
            2026: f"=I{0}-Assumptions!J{ASSUMP_ROWS['buybacks']}",
            2027: f"=J{0}-Assumptions!K{ASSUMP_ROWS['buybacks']}",
            2028: f"=K{0}-Assumptions!L{ASSUMP_ROWS['buybacks']}",
        },
        note="Prior Treasury - Buybacks (more negative)",
    )
    ws.cell(row=ts_row, column=year_col(2026)).value = f"=I{ts_row}-Assumptions!J{ASSUMP_ROWS['buybacks']}"
    ws.cell(row=ts_row, column=year_col(2027)).value = f"=J{ts_row}-Assumptions!K{ASSUMP_ROWS['buybacks']}"
    ws.cell(row=ts_row, column=year_col(2028)).value = f"=K{ts_row}-Assumptions!L{ASSUMP_ROWS['buybacks']}"

    # AOCI
    aoci_row = row_data(
        "AOCI & Other Adjustments",
        data_key="AOCI & Other",
        forecast={
            2026: f"=I{0}",
            2027: f"=J{0}",
            2028: f"=K{0}",
        },
        note="Held constant (unpredictable, immaterial)",
    )
    ws.cell(row=aoci_row, column=year_col(2026)).value = f"=I{aoci_row}"
    ws.cell(row=aoci_row, column=year_col(2027)).value = f"=J{aoci_row}"
    ws.cell(row=aoci_row, column=year_col(2028)).value = f"=K{aoci_row}"

    empty()

    # Total Equity
    te_row = row_data(
        "Total Stockholders' Equity",
        data_key="Total Stockholders' Equity",
        forecast={
            2026: f"=J{csapic_row}+J{re_row}+J{ts_row}+J{aoci_row}",
            2027: f"=K{csapic_row}+K{re_row}+K{ts_row}+K{aoci_row}",
            2028: f"=L{csapic_row}+L{re_row}+L{ts_row}+L{aoci_row}",
        },
        subtotal=True,
        note="APIC + RE + Treasury + AOCI",
    )

    empty()

    # Total L&E
    tle_row = row_data(
        "Total Liabilities & Equity",
        data_key="Total Liabilities & Equity",
        forecast={
            2026: f"=J{tl_row}+J{te_row}",
            2027: f"=K{tl_row}+K{te_row}",
            2028: f"=L{tl_row}+L{te_row}",
        },
        total=True,
        note="Total Liabilities + Total Equity",
    )

    empty()

    # =========================================================================
    # BALANCE CHECK
    # =========================================================================
    section("BALANCE CHECK")

    check_row = get_row()
    c = ws.cell(row=check_row, column=2, value="Difference (Assets - L&E)")
    c.font = Font(name="Arial", size=10, bold=True, color="2E4057")

    for year in YEARS_ACTUAL:
        col = year_col(year)
        ta_val = BS["Total Assets"].get(year, 0)
        tle_val = BS["Total Liabilities & Equity"].get(year, 0)
        diff = ta_val - tle_val
        c = ws.cell(row=check_row, column=col, value=diff)
        c.font = FONT_CHECK_OK if abs(diff) < 1 else FONT_CHECK_FAIL
        c.number_format = FMT_CURRENCY
        c.alignment = ALIGN_RIGHT
        c.fill = FILL_CHECK

    for year in YEARS_FORECAST:
        col = year_col(year)
        cl = col_letter(col)
        c = ws.cell(row=check_row, column=col,
                    value=f"={cl}{ta_row}-{cl}{tle_row}")
        c.font = FONT_CHECK_OK
        c.number_format = FMT_CURRENCY
        c.alignment = ALIGN_RIGHT
        c.fill = FILL_CHECK

    ws.cell(row=check_row, column=COL_NOTE, value="MUST BE $0 — model integrity check").font = Font(
        name="Arial", size=9, bold=True, italic=True, color="FF0000"
    )

    set_row(check_row + 2)

    # Conditional formatting note
    r = get_row()
    ws.cell(row=r, column=2, value="✓ If all zeros → model is balanced").font = FONT_CHECK_OK
    r += 1
    ws.cell(row=r, column=2, value="✗ If non-zero → check formulas for errors").font = FONT_CHECK_FAIL

    # =========================================================================
    # ROW MAP
    # =========================================================================
    r += 3
    ws.cell(row=r, column=2, value="Row Reference Map (for Cash Flow linking):").font = FONT_SMALL
    refs = {
        "Cash": cash_row, "ST Investments": st_inv_row, "AR": ar_row,
        "Customer Accts": cust_row, "Total Current Assets": tca_row,
        "PP&E": ppe_row, "Goodwill": gw_row, "Total Assets": ta_row,
        "AP": ap_row, "Accrued": acc_row, "Total Current Liab": tcl_row,
        "LT Debt": ltd_row, "Total Liabilities": tl_row,
        "APIC": csapic_row, "Retained Earnings": re_row,
        "Treasury": ts_row, "Total Equity": te_row,
        "Total L&E": tle_row, "Balance Check": check_row,
    }
    for i, (name, row_num) in enumerate(refs.items()):
        ws.cell(row=r + 1 + i, column=2, value=f"{name}: Row {row_num}").font = Font(name="Arial", size=8, color="BBBBBB")

    # Freeze panes
    ws.freeze_panes = "C5"

    # Print summary
    print(f"\n  Balance Sheet Row Map:")
    for name, row_num in refs.items():
        print(f"    {name:25s} Row {row_num}")

    return ws


def main():
    print(f"Opening model: {MODEL_PATH}")
    if not os.path.exists(MODEL_PATH):
        print("  ✗ File not found! Run previous scripts first.")
        return

    wb = load_workbook(MODEL_PATH)
    print(f"  ✓ Loaded. Tabs: {', '.join(wb.sheetnames)}")

    build_balance_sheet(wb)

    wb.save(MODEL_PATH)
    print(f"\n  ✓ Balance Sheet tab added and saved.")
    print(f"  ✓ File: {MODEL_PATH}")
    print(f"\n  → Open in Excel and press Ctrl+Shift+F9 to recalculate")
    print(f"  → Check the BALANCE CHECK row — must be $0 for forecast years")
    print(f"  → Cash row is a PLACEHOLDER (yellow) — will link to CF statement next")
    print(f"\n  Next step: Cash Flow Statement tab")


if __name__ == "__main__":
    main()
