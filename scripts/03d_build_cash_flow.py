"""
PayPal (PYPL) - Cash Flow Statement Tab Builder
================================================
Builds the Cash Flow Statement and links Ending Cash back to
the Balance Sheet, closing the 3-statement loop.

Key linkages:
  - Net Income → from Income Statement
  - D&A, SBC, CapEx, Buybacks → from Assumptions
  - Working Capital changes → from Balance Sheet deltas
  - Ending Cash → linked BACK to Balance Sheet Cash row

Run from scripts/ folder AFTER 03c_build_balance_sheet.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model", "PYPL_Financial_Model.xlsx")

# =============================================================================
# STYLES
# =============================================================================
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="2E4057")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_YEAR_FC = Font(name="Arial", size=10, bold=True, color="CCDDFF")
FONT_LABEL = Font(name="Arial", size=10, color="000000")
FONT_LABEL_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
FONT_FORMULA = Font(name="Arial", size=10, color="000000")
FONT_GREEN = Font(name="Arial", size=10, color="008000")
FONT_GREEN_BOLD = Font(name="Arial", size=10, bold=True, color="008000")
FONT_BLUE = Font(name="Arial", size=10, color="0000FF")
FONT_SMALL = Font(name="Arial", size=9, italic=True, color="888888")
FONT_MARGIN = Font(name="Arial", size=9, italic=True, color="666666")
FONT_MARGIN_GREEN = Font(name="Arial", size=9, italic=True, color="008000")

FILL_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_ACTUAL = PatternFill("solid", fgColor="F2F2F2")
FILL_FORECAST = PatternFill("solid", fgColor="E8F0FE")
FILL_SUBTOTAL = PatternFill("solid", fgColor="E8E8E8")
FILL_SUBTOTAL_FC = PatternFill("solid", fgColor="D6E4F0")
FILL_TOTAL = PatternFill("solid", fgColor="D6E4F0")
FILL_INPUT = PatternFill("solid", fgColor="FFFFCC")
FILL_CASH = PatternFill("solid", fgColor="E8FFE8")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=Side(style="medium", color="000000"))
BORDER_THIN = Border(bottom=Side(style="thin", color="CCCCCC"))
BORDER_DOUBLE = Border(bottom=Side(style="double", color="000000"))
BORDER_SECTION = Border(
    top=Side(style="medium", color="2E4057"),
    bottom=Side(style="thin", color="2E4057"),
)

FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'

# =============================================================================
# ROW REFERENCES TO OTHER TABS
# =============================================================================
# Income Statement
IS = {
    "revenue": 7,
    "net_income": 39,
}

# Balance Sheet (from 03c trace)
BS = {
    "cash": 7,
    "st_investments": 8,
    "ar": 9,
    "customer_accts": 10,
    "tca": 12,
    "ppe": 16,
    "goodwill": 17,
    "ta": 21,
    "ap": 26,
    "accrued": 27,
    "tcl": 29,
    "ltd": 33,
    "other_ncl": 34,
    "tl": 36,
    "csapic": 41,
    "re": 42,
    "treasury": 43,
    "te": 46,
    "tle": 48,
}

# Assumptions
ASSUMP = {
    "revenue": 7,
    "capex": 22,
    "capex_pct": 23,
    "da": 24,
    "sbc_pct": 15,
    "buybacks": 34,
    "long_term_debt": 31,
}

# =============================================================================
# HISTORICAL DATA (USD Millions)
# =============================================================================
YEARS_ACTUAL = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
YEARS_FORECAST = [2026, 2027, 2028]

CF = {
    "Net Income": {
        2019: 2459, 2020: 4202, 2021: 4169, 2022: 2419,
        2023: 4246, 2024: 4573, 2025: 4310,
    },
    "Depreciation & Amortization": {
        2019: 810, 2020: 1011, 2021: 1200, 2022: 1274,
        2023: 1290, 2024: 1259, 2025: 1280,
    },
    "Stock-Based Compensation": {
        2019: 1021, 2020: 1378, 2021: 1510, 2022: 1456,
        2023: 1390, 2024: 1399, 2025: 1350,
    },
    "Other Operating Adjustments": {
        2019: 272, 2020: -738, 2021: -536, 2022: 664,
        2023: -2083, 2024: -831, 2025: 180,
    },
    "Cash from Operations": {
        2019: 4562, 2020: 5853, 2021: 6343, 2022: 5813,
        2023: 4843, 2024: 6400, 2025: 7120,
    },
    "Capital Expenditures": {
        2019: -704, 2020: -866, 2021: -908, 2022: -706,
        2023: -596, 2024: -516, 2025: -505,
    },
    "Other Investing Activities": {
        2019: -1871, 2020: -7030, 2021: -3087, 2022: -1050,
        2023: -1200, 2024: 800, 2025: -500,
    },
    "Cash from Investing": {
        2019: -2575, 2020: -7896, 2021: -3995, 2022: -1756,
        2023: -1796, 2024: 284, 2025: -1005,
    },
    "Debt Issuance / (Repayment)": {
        2019: 1996, 2020: 3994, 2021: -990, 2022: 2398,
        2023: -741, 2024: 203, 2025: 108,
    },
    "Share Repurchases": {
        2019: -3333, 2020: -1021, 2021: -3397, 2022: -4200,
        2023: -5100, 2024: -5400, 2025: -6000,
    },
    "Other Financing Activities": {
        2019: -308, 2020: -350, 2021: -240, 2022: -100,
        2023: -230, 2024: -150, 2025: -200,
    },
    "Cash from Financing": {
        2019: -1645, 2020: 2623, 2021: -4627, 2022: -1902,
        2023: -6071, 2024: -5347, 2025: -6092,
    },
    "Net Change in Cash": {
        2019: 342, 2020: 580, 2021: -2279, 2022: 4155,
        2023: -3024, 2024: 1337, 2025: 23,
    },
    "Beginning Cash": {
        2019: 7007, 2020: 7349, 2021: 4794, 2022: 5197,
        2023: 7776, 2024: 9081, 2025: 6662,
    },
    "Ending Cash": {
        2019: 7349, 2020: 4794, 2021: 5197, 2022: 7776,
        2023: 9081, 2024: 6662, 2025: 8049,
    },
}


def year_col(year):
    if year in YEARS_ACTUAL:
        return 3 + YEARS_ACTUAL.index(year)
    return 10 + YEARS_FORECAST.index(year)


def cl(col_num):
    return chr(64 + col_num)


# =============================================================================
# BUILD CASH FLOW STATEMENT
# =============================================================================
def build_cash_flow(wb):
    if "Cash Flow" in wb.sheetnames:
        del wb["Cash Flow"]

    bs_idx = wb.sheetnames.index("Balance Sheet")
    ws = wb.create_sheet("Cash Flow", bs_idx + 1)
    ws.sheet_properties.tabColor = "FFC107"

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    for c in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[c].width = 14
    ws.column_dimensions["M"].width = 38

    COL_NOTE = 13
    row = [0]

    def sr(r): row[0] = r
    def gr(): return row[0]

    # --- Title ---
    sr(1)
    ws.merge_cells("B1:L1")
    ws.cell(row=1, column=2, value="PayPal (PYPL) — Consolidated Cash Flow Statement").font = FONT_TITLE
    ws.cell(row=2, column=2, value="USD Millions ($M)").font = FONT_SMALL
    ws.cell(row=2, column=COL_NOTE, value="Notes").font = Font(name="Arial", size=10, bold=True, italic=True, color="888888")

    # --- Year Headers ---
    ws.cell(row=3, column=2).fill = FILL_HEADER
    ws.cell(row=3, column=COL_NOTE).fill = FILL_HEADER
    for year in YEARS_ACTUAL:
        c = ws.cell(row=3, column=year_col(year), value=str(year))
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER; c.number_format = '@'
    for year in YEARS_FORECAST:
        c = ws.cell(row=3, column=year_col(year), value=f"{year}E")
        c.font = FONT_YEAR_FC; c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER

    for year in YEARS_ACTUAL:
        c = ws.cell(row=4, column=year_col(year), value="Actual")
        c.font = Font(name="Arial", size=8, italic=True, color="888888"); c.alignment = ALIGN_CENTER
    for year in YEARS_FORECAST:
        c = ws.cell(row=4, column=year_col(year), value="Forecast")
        c.font = Font(name="Arial", size=8, italic=True, color="0000FF"); c.alignment = ALIGN_CENTER

    sr(5)

    # =========================================================================
    # HELPERS
    # =========================================================================
    def section(title):
        r = gr() + 1
        ws.merge_cells(f"B{r}:L{r}")
        c = ws.cell(row=r, column=2, value=title)
        c.font = Font(name="Arial", size=11, bold=True, color="2E4057")
        c.fill = FILL_SECTION; c.border = BORDER_SECTION
        for col in range(3, 14):
            ws.cell(row=r, column=col).fill = FILL_SECTION
            ws.cell(row=r, column=col).border = BORDER_SECTION
        sr(r + 1)

    def data_row(label, data_key=None, forecast=None, note="",
                 subtotal=False, total=False, is_cash=False):
        r = gr()

        c = ws.cell(row=r, column=2, value=label)
        c.font = FONT_LABEL_BOLD if (subtotal or total) else FONT_LABEL
        c.alignment = ALIGN_LEFT

        border = BORDER_DOUBLE if total else (BORDER_BOTTOM if subtotal else BORDER_THIN)

        # Actuals
        if data_key and data_key in CF:
            for year in YEARS_ACTUAL:
                val = CF[data_key].get(year)
                if val is not None:
                    c = ws.cell(row=r, column=year_col(year), value=val)
                    c.font = Font(name="Arial", size=10, bold=subtotal or total, color="000000")
                    c.number_format = FMT_CURRENCY
                    c.alignment = ALIGN_RIGHT
                    if is_cash:
                        c.fill = FILL_CASH
                    elif subtotal or total:
                        c.fill = FILL_SUBTOTAL
                    else:
                        c.fill = FILL_ACTUAL
                    c.border = border

        # Forecast
        if forecast:
            for year in YEARS_FORECAST:
                f = forecast.get(year)
                if f is not None:
                    c = ws.cell(row=r, column=year_col(year), value=f)
                    if subtotal or total:
                        c.font = FONT_GREEN_BOLD
                    else:
                        c.font = FONT_GREEN
                    c.number_format = FMT_CURRENCY
                    c.alignment = ALIGN_RIGHT
                    if is_cash:
                        c.fill = FILL_CASH
                    elif subtotal or total:
                        c.fill = FILL_SUBTOTAL_FC
                    else:
                        c.fill = FILL_FORECAST
                    c.border = border

        if note:
            ws.cell(row=r, column=COL_NOTE, value=note).font = FONT_SMALL

        sr(r + 1)
        return r

    def margin_row(label, data, forecast=None):
        r = gr()
        ws.cell(row=r, column=2, value=label).font = FONT_MARGIN
        for year in YEARS_ACTUAL:
            val = data.get(year)
            if val is not None:
                c = ws.cell(row=r, column=year_col(year), value=val)
                c.font = FONT_MARGIN; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
        if forecast:
            for year in YEARS_FORECAST:
                f = forecast.get(year)
                if f:
                    c = ws.cell(row=r, column=year_col(year), value=f)
                    c.font = FONT_MARGIN_GREEN; c.number_format = FMT_PCT; c.alignment = ALIGN_RIGHT
        sr(r + 1)

    def empty():
        sr(gr() + 1)

    # =========================================================================
    # OPERATING ACTIVITIES
    # =========================================================================
    section("CASH FROM OPERATING ACTIVITIES")

    # Net Income (from IS)
    ni_row = data_row(
        "Net Income",
        data_key="Net Income",
        forecast={
            2026: f"='Income Statement'!J{IS['net_income']}",
            2027: f"='Income Statement'!K{IS['net_income']}",
            2028: f"='Income Statement'!L{IS['net_income']}",
        },
        note="From Income Statement (green = cross-sheet link)",
    )

    empty()

    # Adjustments for non-cash items
    ws.cell(row=gr(), column=2, value="Adjustments for non-cash items:").font = Font(name="Arial", size=9, italic=True, color="666666")
    sr(gr() + 1)

    # D&A
    da_row = data_row(
        "  (+) Depreciation & Amortization",
        data_key="Depreciation & Amortization",
        forecast={
            2026: f"=Assumptions!J{ASSUMP['da']}",
            2027: f"=Assumptions!K{ASSUMP['da']}",
            2028: f"=Assumptions!L{ASSUMP['da']}",
        },
        note="From Assumptions tab (non-cash add-back)",
    )

    # SBC
    sbc_row = data_row(
        "  (+) Stock-Based Compensation",
        data_key="Stock-Based Compensation",
        forecast={
            2026: f"=Assumptions!J{ASSUMP['revenue']}*Assumptions!J{ASSUMP['sbc_pct']}",
            2027: f"=Assumptions!K{ASSUMP['revenue']}*Assumptions!K{ASSUMP['sbc_pct']}",
            2028: f"=Assumptions!L{ASSUMP['revenue']}*Assumptions!L{ASSUMP['sbc_pct']}",
        },
        note="Revenue x SBC % (non-cash add-back)",
    )

    empty()

    # Working Capital Changes
    ws.cell(row=gr(), column=2, value="Changes in working capital:").font = Font(name="Arial", size=9, italic=True, color="666666")
    sr(gr() + 1)

    # Change in AR: -(Current - Prior)
    ar_chg_row = data_row(
        "  (Inc)/Dec in Accounts Receivable",
        forecast={
            2026: f"=-('Balance Sheet'!J{BS['ar']}-'Balance Sheet'!I{BS['ar']})",
            2027: f"=-('Balance Sheet'!K{BS['ar']}-'Balance Sheet'!J{BS['ar']})",
            2028: f"=-('Balance Sheet'!L{BS['ar']}-'Balance Sheet'!K{BS['ar']})",
        },
        note="Negative when AR increases (cash tied up)",
    )
    # Fill actuals for AR change
    ar_data = {
        2020: -(577 - 435), 2021: -(12723 - 577), 2022: -(2201 - 12723),
        2023: -(1882 - 2201), 2024: -(1581 - 1882), 2025: -(1909 - 1581),
    }
    for year, val in ar_data.items():
        c = ws.cell(row=ar_chg_row, column=year_col(year), value=val)
        c.font = FONT_FORMULA; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_ACTUAL

    # Change in Customer Accounts
    cust_chg_row = data_row(
        "  (Inc)/Dec in Customer Accounts",
        forecast={
            2026: f"=-('Balance Sheet'!J{BS['customer_accts']}-'Balance Sheet'!I{BS['customer_accts']})",
            2027: f"=-('Balance Sheet'!K{BS['customer_accts']}-'Balance Sheet'!J{BS['customer_accts']})",
            2028: f"=-('Balance Sheet'!L{BS['customer_accts']}-'Balance Sheet'!K{BS['customer_accts']})",
        },
        note="Large PayPal-specific item (customer funds)",
    )
    cust_data = {
        2020: -(37335 - 27299), 2021: -(0 - 37335), 2022: -(44372 - 0),
        2023: -(46630 - 44372), 2024: -(45702 - 46630), 2025: -(47428 - 45702),
    }
    for year, val in cust_data.items():
        c = ws.cell(row=cust_chg_row, column=year_col(year), value=val)
        c.font = FONT_FORMULA; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_ACTUAL

    # Change in AP
    ap_chg_row = data_row(
        "  Inc/(Dec) in Accounts Payable",
        forecast={
            2026: f"='Balance Sheet'!J{BS['ap']}-'Balance Sheet'!I{BS['ap']}",
            2027: f"='Balance Sheet'!K{BS['ap']}-'Balance Sheet'!J{BS['ap']}",
            2028: f"='Balance Sheet'!L{BS['ap']}-'Balance Sheet'!K{BS['ap']}",
        },
        note="Positive when AP increases (cash preserved)",
    )
    ap_data = {
        2020: 252 - 232, 2021: 197 - 252, 2022: 40140 - 197,
        2023: 42074 - 40140, 2024: 39898 - 42074, 2025: 40438 - 39898,
    }
    for year, val in ap_data.items():
        c = ws.cell(row=ap_chg_row, column=year_col(year), value=val)
        c.font = FONT_FORMULA; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_ACTUAL

    # Change in Accrued
    acc_chg_row = data_row(
        "  Inc/(Dec) in Accrued Expenses",
        forecast={
            2026: f"='Balance Sheet'!J{BS['accrued']}-'Balance Sheet'!I{BS['accrued']}",
            2027: f"='Balance Sheet'!K{BS['accrued']}-'Balance Sheet'!J{BS['accrued']}",
            2028: f"='Balance Sheet'!L{BS['accrued']}-'Balance Sheet'!K{BS['accrued']}",
        },
        note="Positive when accrued liabilities increase",
    )
    acc_data = {
        2020: 38195 - 26687, 2021: 42832 - 38195, 2022: 4868 - 42832,
        2023: 6392 - 4868, 2024: 5592 - 6392, 2025: 6005 - 5592,
    }
    for year, val in acc_data.items():
        c = ws.cell(row=acc_chg_row, column=year_col(year), value=val)
        c.font = FONT_FORMULA; c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_ACTUAL

    # Other operating adjustments (plug to match historical CFO)
    other_op_row = data_row(
        "  Other Operating Adjustments",
        data_key="Other Operating Adjustments",
        forecast={
            2026: "=0",
            2027: "=0",
            2028: "=0",
        },
        note="Conservatively zero for forecast",
    )

    empty()

    # --- CASH FROM OPERATIONS ---
    cfo_row = data_row(
        "Cash from Operations",
        data_key="Cash from Operations",
        forecast={
            2026: f"=J{ni_row}+J{da_row}+J{sbc_row}+J{ar_chg_row}+J{cust_chg_row}+J{ap_chg_row}+J{acc_chg_row}+J{other_op_row}",
            2027: f"=K{ni_row}+K{da_row}+K{sbc_row}+K{ar_chg_row}+K{cust_chg_row}+K{ap_chg_row}+K{acc_chg_row}+K{other_op_row}",
            2028: f"=L{ni_row}+L{da_row}+L{sbc_row}+L{ar_chg_row}+L{cust_chg_row}+L{ap_chg_row}+L{acc_chg_row}+L{other_op_row}",
        },
        subtotal=True,
        note="NI + Non-cash + Working Capital Changes",
    )

    # CFO margin
    cfo_margin = {}
    rev = {2019: 17772, 2020: 21454, 2021: 25371, 2022: 27518, 2023: 29771, 2024: 31797, 2025: 33172}
    for year in YEARS_ACTUAL:
        cfo_val = CF["Cash from Operations"].get(year)
        r_val = rev.get(year)
        if cfo_val and r_val:
            cfo_margin[year] = cfo_val / r_val
    margin_row(
        "  CFO Margin (% of Revenue)",
        cfo_margin,
        forecast={
            2026: f"=J{cfo_row}/'Income Statement'!J{IS['revenue']}",
            2027: f"=K{cfo_row}/'Income Statement'!K{IS['revenue']}",
            2028: f"=L{cfo_row}/'Income Statement'!L{IS['revenue']}",
        },
    )

    empty()

    # =========================================================================
    # INVESTING ACTIVITIES
    # =========================================================================
    section("CASH FROM INVESTING ACTIVITIES")

    capex_row = data_row(
        "Capital Expenditures",
        data_key="Capital Expenditures",
        forecast={
            2026: f"=-Assumptions!J{ASSUMP['capex']}",
            2027: f"=-Assumptions!K{ASSUMP['capex']}",
            2028: f"=-Assumptions!L{ASSUMP['capex']}",
        },
        note="From Assumptions (shown as negative)",
    )

    other_inv_row = data_row(
        "Other Investing Activities",
        data_key="Other Investing Activities",
        forecast={
            2026: "=0",
            2027: "=0",
            2028: "=0",
        },
        note="Conservatively zero (no acquisitions assumed)",
    )

    empty()

    cfi_row = data_row(
        "Cash from Investing",
        data_key="Cash from Investing",
        forecast={
            2026: f"=J{capex_row}+J{other_inv_row}",
            2027: f"=K{capex_row}+K{other_inv_row}",
            2028: f"=L{capex_row}+L{other_inv_row}",
        },
        subtotal=True,
        note="CapEx + Other Investing",
    )

    empty()

    # =========================================================================
    # FINANCING ACTIVITIES
    # =========================================================================
    section("CASH FROM FINANCING ACTIVITIES")

    # Debt change = Current LT Debt - Prior LT Debt (from BS)
    debt_chg_row = data_row(
        "Debt Issuance / (Repayment)",
        data_key="Debt Issuance / (Repayment)",
        forecast={
            2026: f"='Balance Sheet'!J{BS['ltd']}-'Balance Sheet'!I{BS['ltd']}",
            2027: f"='Balance Sheet'!K{BS['ltd']}-'Balance Sheet'!J{BS['ltd']}",
            2028: f"='Balance Sheet'!L{BS['ltd']}-'Balance Sheet'!K{BS['ltd']}",
        },
        note="Change in LT Debt from Balance Sheet",
    )

    buyback_row = data_row(
        "Share Repurchases",
        data_key="Share Repurchases",
        forecast={
            2026: f"=-Assumptions!J{ASSUMP['buybacks']}",
            2027: f"=-Assumptions!K{ASSUMP['buybacks']}",
            2028: f"=-Assumptions!L{ASSUMP['buybacks']}",
        },
        note="From Assumptions (shown as negative cash outflow)",
    )

    other_fin_row = data_row(
        "Other Financing Activities",
        data_key="Other Financing Activities",
        forecast={
            2026: "=0",
            2027: "=0",
            2028: "=0",
        },
        note="Set to zero (no BS counterpart; immaterial)",
    )

    empty()

    cff_row = data_row(
        "Cash from Financing",
        data_key="Cash from Financing",
        forecast={
            2026: f"=J{debt_chg_row}+J{buyback_row}+J{other_fin_row}",
            2027: f"=K{debt_chg_row}+K{buyback_row}+K{other_fin_row}",
            2028: f"=L{debt_chg_row}+L{buyback_row}+L{other_fin_row}",
        },
        subtotal=True,
        note="Debt changes + Buybacks + Other",
    )

    empty()
    empty()

    # =========================================================================
    # NET CHANGE IN CASH & ENDING BALANCE
    # =========================================================================
    section("CASH SUMMARY")

    net_chg_row = data_row(
        "Net Change in Cash",
        data_key="Net Change in Cash",
        forecast={
            2026: f"=J{cfo_row}+J{cfi_row}+J{cff_row}",
            2027: f"=K{cfo_row}+K{cfi_row}+K{cff_row}",
            2028: f"=L{cfo_row}+L{cfi_row}+L{cff_row}",
        },
        subtotal=True,
        note="CFO + CFI + CFF",
    )

    empty()

    # Beginning Cash
    beg_cash_row = data_row(
        "Beginning Cash Balance",
        data_key="Beginning Cash",
        forecast={
            # 2026 beginning = 2025 ending cash (from BS)
            2026: f"='Balance Sheet'!I{BS['cash']}",
            2027: f"=J{0}",  # placeholder, fix below
            2028: f"=K{0}",
        },
        is_cash=True,
        note="Prior year ending cash balance",
    )

    # Ending Cash = Beginning + Net Change
    end_cash_row = data_row(
        "Ending Cash Balance",
        data_key="Ending Cash",
        forecast={
            2026: f"=J{beg_cash_row}+J{net_chg_row}",
            2027: f"=K{beg_cash_row}+K{net_chg_row}",
            2028: f"=L{beg_cash_row}+L{net_chg_row}",
        },
        total=True, is_cash=True,
        note="THIS LINKS BACK TO BALANCE SHEET CASH ROW",
    )

    # Fix beginning cash for 2027/2028 to reference prior year ending cash
    ws.cell(row=beg_cash_row, column=year_col(2027)).value = f"=J{end_cash_row}"
    ws.cell(row=beg_cash_row, column=year_col(2028)).value = f"=K{end_cash_row}"

    empty()
    empty()

    # =========================================================================
    # FREE CASH FLOW (Key metric for DCF)
    # =========================================================================
    section("FREE CASH FLOW (for DCF)")

    fcf_row = data_row(
        "Free Cash Flow (CFO - CapEx)",
        forecast={
            2026: f"=J{cfo_row}+J{capex_row}",
            2027: f"=K{cfo_row}+K{capex_row}",
            2028: f"=L{cfo_row}+L{capex_row}",
        },
        subtotal=True,
        note="Key input for DCF valuation",
    )
    # Fill FCF actuals
    for year in YEARS_ACTUAL:
        cfo_val = CF["Cash from Operations"].get(year, 0)
        capex_val = CF["Capital Expenditures"].get(year, 0)
        fcf_val = cfo_val + capex_val  # capex is already negative
        c = ws.cell(row=fcf_row, column=year_col(year), value=fcf_val)
        c.font = Font(name="Arial", size=10, bold=True, color="000000")
        c.number_format = FMT_CURRENCY; c.alignment = ALIGN_RIGHT; c.fill = FILL_SUBTOTAL
        c.border = BORDER_BOTTOM

    # FCF margin
    fcf_margin = {}
    for year in YEARS_ACTUAL:
        cfo_val = CF["Cash from Operations"].get(year, 0)
        capex_val = CF["Capital Expenditures"].get(year, 0)
        r_val = rev.get(year)
        if r_val:
            fcf_margin[year] = (cfo_val + capex_val) / r_val
    margin_row(
        "  FCF Margin (% of Revenue)",
        fcf_margin,
        forecast={
            2026: f"=J{fcf_row}/'Income Statement'!J{IS['revenue']}",
            2027: f"=K{fcf_row}/'Income Statement'!K{IS['revenue']}",
            2028: f"=L{fcf_row}/'Income Statement'!L{IS['revenue']}",
        },
    )

    # FCF per share
    fcf_ps_row = data_row(
        "FCF per Share",
        forecast={
            2026: f"=J{fcf_row}/'Income Statement'!J45",
            2027: f"=K{fcf_row}/'Income Statement'!K45",
            2028: f"=L{fcf_row}/'Income Statement'!L45",
        },
        note="FCF / Diluted Shares (from IS)",
    )
    # Fill actuals
    shares = {2019: 1188, 2020: 1187, 2021: 1183, 2022: 1156, 2023: 1105, 2024: 1053, 2025: 941}
    for year in YEARS_ACTUAL:
        cfo_val = CF["Cash from Operations"].get(year, 0)
        capex_val = CF["Capital Expenditures"].get(year, 0)
        s = shares.get(year)
        if s:
            c = ws.cell(row=fcf_ps_row, column=year_col(year), value=round((cfo_val + capex_val) / s, 2))
            c.font = FONT_FORMULA; c.number_format = '#,##0.00'; c.alignment = ALIGN_RIGHT; c.fill = FILL_ACTUAL

    # =========================================================================
    # ROW MAP
    # =========================================================================
    empty()
    empty()
    r = gr()
    ws.cell(row=r, column=2, value="Row Reference Map (for DCF linking):").font = FONT_SMALL
    refs = {
        "Net Income": ni_row, "D&A": da_row, "SBC": sbc_row,
        "CFO": cfo_row, "CapEx": capex_row, "CFI": cfi_row,
        "Buybacks": buyback_row, "CFF": cff_row,
        "Net Change": net_chg_row, "Beg Cash": beg_cash_row,
        "End Cash": end_cash_row, "FCF": fcf_row,
    }
    for i, (name, rn) in enumerate(refs.items()):
        ws.cell(row=r + 1 + i, column=2, value=f"{name}: Row {rn}").font = Font(name="Arial", size=8, color="BBBBBB")

    ws.freeze_panes = "C5"

    print(f"\n  Cash Flow Statement Row Map:")
    for name, rn in refs.items():
        print(f"    {name:20s} Row {rn}")

    return ws, end_cash_row


# =============================================================================
# LINK CASH BACK TO BALANCE SHEET
# =============================================================================
def link_cash_to_balance_sheet(wb, cf_end_cash_row):
    """Replace the Cash placeholder on BS with a formula linking to CF Ending Cash."""
    ws = wb["Balance Sheet"]

    cash_row = BS["cash"]  # Row 7

    for year in YEARS_FORECAST:
        col = year_col(year)
        c_letter = cl(col)

        # Link BS Cash = CF Ending Cash
        formula = f"='Cash Flow'!{c_letter}{cf_end_cash_row}"
        c = ws.cell(row=cash_row, column=col, value=formula)
        c.font = FONT_GREEN_BOLD
        c.number_format = FMT_CURRENCY
        c.alignment = ALIGN_RIGHT
        c.fill = FILL_CASH

    # Update the note
    ws.cell(row=cash_row, column=13, value="LINKED to Cash Flow Ending Balance (model is closed)").font = Font(
        name="Arial", size=9, bold=True, italic=True, color="008000"
    )

    print(f"\n  ✓ Balance Sheet Cash (Row {cash_row}) now linked to Cash Flow Ending Cash (Row {cf_end_cash_row})")
    print(f"    Formulas: ='Cash Flow'!J{cf_end_cash_row}, K{cf_end_cash_row}, L{cf_end_cash_row}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"Opening model: {MODEL_PATH}")
    if not os.path.exists(MODEL_PATH):
        print("  File not found!")
        return

    wb = load_workbook(MODEL_PATH)
    print(f"  Loaded. Tabs: {', '.join(wb.sheetnames)}")

    ws, end_cash_row = build_cash_flow(wb)

    link_cash_to_balance_sheet(wb, end_cash_row)

    wb.save(MODEL_PATH)
    print(f"\n  ✓ Cash Flow tab added.")
    print(f"  ✓ Balance Sheet Cash linked to Cash Flow Ending Cash.")
    print(f"  ✓ File: {MODEL_PATH}")
    print(f"\n  → Open Excel, press Ctrl+Shift+F9")
    print(f"  → Check Balance Sheet 'BALANCE CHECK' row — should now be $0")
    print(f"  → The 3-statement model loop is CLOSED.")
    print(f"\n  Next: Ratios tab + DCF tab")


if __name__ == "__main__":
    main()
