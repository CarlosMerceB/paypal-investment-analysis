"""
PayPal (PYPL) - Income Statement Tab Builder
=============================================
Opens existing PYPL_Financial_Model.xlsx and adds the Income Statement tab.

Historical actuals (FY2019-2025): hardcoded values from extraction.
Forecast (FY2026E-2028E): Excel formulas referencing Assumptions tab.

Run from scripts/ folder AFTER 03_build_excel_model.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# =============================================================================
# CONFIGURATION
# =============================================================================
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "model", "PYPL_Financial_Model.xlsx")

# =============================================================================
# STYLE DEFINITIONS (same as main model)
# =============================================================================
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="2E4057")
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="2E4057")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_LABEL = Font(name="Arial", size=10, color="000000")
FONT_LABEL_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
FONT_INPUT_BLUE = Font(name="Arial", size=10, color="0000FF")
FONT_FORMULA = Font(name="Arial", size=10, color="000000")
FONT_GREEN = Font(name="Arial", size=10, color="008000")
FONT_GREEN_BOLD = Font(name="Arial", size=10, bold=True, color="008000")
FONT_SUBTOTAL = Font(name="Arial", size=10, bold=True, color="000000")
FONT_SUBTOTAL_GREEN = Font(name="Arial", size=10, bold=True, color="008000")
FONT_SMALL_NOTE = Font(name="Arial", size=9, italic=True, color="888888")
FONT_YEAR_FORECAST = Font(name="Arial", size=10, bold=True, color="CCDDFF")

FILL_HEADER = PatternFill("solid", fgColor="2E4057")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_ACTUAL = PatternFill("solid", fgColor="F2F2F2")
FILL_FORECAST = PatternFill("solid", fgColor="E8F0FE")
FILL_SUBTOTAL = PatternFill("solid", fgColor="E8E8E8")
FILL_SUBTOTAL_FC = PatternFill("solid", fgColor="D6E4F0")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=Side(style="medium", color="000000"))
BORDER_THIN = Border(bottom=Side(style="thin", color="CCCCCC"))
BORDER_SECTION = Border(
    top=Side(style="medium", color="2E4057"),
    bottom=Side(style="thin", color="2E4057"),
)
BORDER_DOUBLE = Border(bottom=Side(style="double", color="000000"))

FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_PER_SHARE = '#,##0.00;(#,##0.00);"-"'
FMT_SHARES = '#,##0'

# =============================================================================
# ASSUMPTIONS TAB ROW REFERENCES
# =============================================================================
# These must match the exact row positions in the Assumptions tab.
# Column mapping: C=2019, D=2020, E=2021, F=2022, G=2023, H=2024, I=2025
#                 J=2026E, K=2027E, L=2028E

ASSUMP = {
    "revenue": 7,           # Total Revenue ($M)
    "revenue_growth": 8,    # Revenue Growth (%)
    "gross_margin": 9,      # Gross Margin (%)
    "total_opex_pct": 12,   # Total OpEx (% of Revenue)
    "operating_margin": 13, # Operating Margin (%)
    "da_pct": 14,           # D&A (% of Revenue)
    "sbc_pct": 15,          # SBC (% of Revenue)
    "tax_rate": 18,         # Effective Tax Rate (%)
    "interest_expense": 19, # Interest Expense ($M)
}

# =============================================================================
# HISTORICAL DATA (USD Millions — from extraction)
# =============================================================================
YEARS_ACTUAL = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
YEARS_FORECAST = [2026, 2027, 2028]

# Income Statement actuals
IS_DATA = {
    "Total Revenue": {
        2019: 17772, 2020: 21454, 2021: 25371, 2022: 27518,
        2023: 29771, 2024: 31797, 2025: 33172,
    },
    "Cost of Revenue": {
        2019: 10004, 2020: 12253, 2021: 14233, 2022: 15326,
        2023: 16395, 2024: 17141, 2025: 17862,
    },
    "Gross Profit": {
        2019: 7768, 2020: 9201, 2021: 11138, 2022: 12192,
        2023: 13376, 2024: 14656, 2025: 15310,
    },
    "Operating Expenses": {
        2019: 5101, 2020: 5912, 2021: 6936, 2022: 8056,
        2023: 8348, 2024: 8938, 2025: 9302,
    },
    "Operating Income": {
        2019: 2667, 2020: 3289, 2021: 4202, 2022: 4136,
        2023: 5028, 2024: 5718, 2025: 6008,
    },
    "Interest Income": {
        2019: 93, 2020: 48, 2021: 23, 2022: 150,
        2023: 482, 2024: 532, 2025: 498,
    },
    "Interest Expense": {
        2019: -234, 2020: -264, 2021: -232, 2022: -309,
        2023: -365, 2024: -392, 2025: -398,
    },
    "Other Income (Expense)": {
        2019: 120, 2020: 1776, 2021: -16, 2022: -494,
        2023: -64, 2024: 10, 2025: -271,
    },
    "Income Before Taxes": {
        2019: 2646, 2020: 4849, 2021: 3977, 2022: 3483,
        2023: 5081, 2024: 5868, 2025: 5837,
    },
    "Income Tax Expense": {
        2019: 187, 2020: 647, 2021: -192, 2022: 1064,
        2023: 835, 2024: 1295, 2025: 1527,
    },
    "Net Income": {
        2019: 2459, 2020: 4202, 2021: 4169, 2022: 2419,
        2023: 4246, 2024: 4573, 2025: 4310,
    },
    "Diluted EPS": {
        2019: 2.07, 2020: 3.54, 2021: 3.52, 2022: 2.09,
        2023: 3.84, 2024: 4.34, 2025: 4.58,
    },
    "Diluted Shares (M)": {
        2019: 1188, 2020: 1187, 2021: 1183, 2022: 1156,
        2023: 1105, 2024: 1053, 2025: 941,
    },
}

# Growth rates and margins (calculated for reference row)
def calc_growth(data, year, prev_year):
    if prev_year in data and year in data and data[prev_year] != 0:
        return (data[year] - data[prev_year]) / abs(data[prev_year])
    return None

def calc_margin(numerator_data, denominator_data, year):
    if year in numerator_data and year in denominator_data and denominator_data[year] != 0:
        return numerator_data[year] / denominator_data[year]
    return None


# =============================================================================
# INCOME STATEMENT COLUMN LAYOUT
# =============================================================================
# Same as Assumptions: B=label, C-I=actuals, J-L=forecast, M=notes
COL_LABEL = 2       # B
COL_FIRST_YEAR = 3  # C = 2019
COL_LAST_ACTUAL = 9 # I = 2025
COL_FIRST_FC = 10   # J = 2026E
COL_LAST_FC = 12    # L = 2028E
COL_NOTE = 13       # M

def year_col(year):
    if year in YEARS_ACTUAL:
        return COL_FIRST_YEAR + YEARS_ACTUAL.index(year)
    return COL_FIRST_FC + YEARS_FORECAST.index(year)

def fc_col_letter(year):
    """Get Excel column letter for a forecast year."""
    return chr(64 + year_col(year))  # 10→J, 11→K, 12→L

def act_col_letter(year):
    """Get Excel column letter for an actual year."""
    return chr(64 + year_col(year))


# =============================================================================
# BUILD INCOME STATEMENT
# =============================================================================
def build_income_statement(wb):
    # Remove placeholder sheet if exists
    if "Income Statement" in wb.sheetnames:
        del wb["Income Statement"]

    ws = wb.create_sheet("Income Statement", 2)  # Position after Assumptions
    ws.sheet_properties.tabColor = "28A745"

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["M"].width = 30

    # =========================================================================
    # TITLE
    # =========================================================================
    row = 1
    ws.merge_cells(f"B{row}:L{row}")
    ws.cell(row=row, column=2, value="PayPal (PYPL) — Consolidated Income Statement").font = FONT_TITLE

    row = 2
    ws.cell(row=row, column=2, value="USD Millions ($M) except per-share data").font = FONT_SMALL_NOTE
    ws.cell(row=row, column=COL_NOTE, value="Notes").font = Font(name="Arial", size=10, bold=True, italic=True, color="888888")

    # =========================================================================
    # YEAR HEADERS (Row 3)
    # =========================================================================
    row = 3
    ws.cell(row=row, column=2).fill = FILL_HEADER
    ws.cell(row=row, column=COL_NOTE).fill = FILL_HEADER

    for year in YEARS_ACTUAL:
        c = ws.cell(row=row, column=year_col(year), value=str(year))
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.number_format = '@'

    for year in YEARS_FORECAST:
        c = ws.cell(row=row, column=year_col(year), value=f"{year}E")
        c.font = FONT_YEAR_FORECAST
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

    # Period type row
    row = 4
    for year in YEARS_ACTUAL:
        c = ws.cell(row=row, column=year_col(year), value="Actual")
        c.font = Font(name="Arial", size=8, italic=True, color="888888")
        c.alignment = ALIGN_CENTER
    for year in YEARS_FORECAST:
        c = ws.cell(row=row, column=year_col(year), value="Forecast")
        c.font = Font(name="Arial", size=8, italic=True, color="0000FF")
        c.alignment = ALIGN_CENTER

    # =========================================================================
    # HELPERS
    # =========================================================================
    current_row = 5

    def section_header(title):
        nonlocal current_row
        current_row += 1
        ws.merge_cells(f"B{current_row}:L{current_row}")
        c = ws.cell(row=current_row, column=2, value=title)
        c.font = Font(name="Arial", size=11, bold=True, color="2E4057")
        c.fill = FILL_SECTION
        c.border = BORDER_SECTION
        for col in range(3, 14):
            ws.cell(row=current_row, column=col).fill = FILL_SECTION
            ws.cell(row=current_row, column=col).border = BORDER_SECTION
        current_row += 1
        return current_row

    def data_row(label, actuals_key=None, forecast_formulas=None,
                 fmt=FMT_CURRENCY, note="", is_subtotal=False, is_total=False,
                 margin_data=None, margin_fmt=FMT_PCT):
        """
        Add a data row.
        actuals_key: key in IS_DATA for historical values
        forecast_formulas: dict {year: "=formula_string"} for forecast columns
        margin_data: optional dict of calculated margins for actual years
        """
        nonlocal current_row

        # Label
        c = ws.cell(row=current_row, column=COL_LABEL, value=label)
        if is_total:
            c.font = Font(name="Arial", size=10, bold=True, color="000000")
        elif is_subtotal:
            c.font = FONT_LABEL_BOLD
        else:
            c.font = FONT_LABEL
        c.alignment = ALIGN_LEFT

        border = BORDER_DOUBLE if is_total else (BORDER_BOTTOM if is_subtotal else BORDER_THIN)

        # Actual values
        if actuals_key and actuals_key in IS_DATA:
            for year in YEARS_ACTUAL:
                col = year_col(year)
                val = IS_DATA[actuals_key].get(year)
                if val is not None:
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.font = FONT_SUBTOTAL if (is_subtotal or is_total) else FONT_FORMULA
                    c.number_format = fmt
                    c.alignment = ALIGN_RIGHT
                    c.fill = FILL_SUBTOTAL if is_subtotal else (FILL_SUBTOTAL if is_total else FILL_ACTUAL)
                    c.border = border

        elif margin_data:
            for year in YEARS_ACTUAL:
                col = year_col(year)
                val = margin_data.get(year)
                if val is not None:
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.font = Font(name="Arial", size=9, italic=True, color="666666")
                    c.number_format = margin_fmt
                    c.alignment = ALIGN_RIGHT
                    c.fill = FILL_ACTUAL

        # Forecast formulas (green font = cross-sheet link)
        if forecast_formulas:
            for year in YEARS_FORECAST:
                col = year_col(year)
                formula = forecast_formulas.get(year)
                if formula:
                    c = ws.cell(row=current_row, column=col, value=formula)
                    if is_subtotal or is_total:
                        c.font = FONT_SUBTOTAL_GREEN
                    else:
                        c.font = FONT_GREEN
                    c.number_format = fmt
                    c.alignment = ALIGN_RIGHT
                    c.fill = FILL_SUBTOTAL_FC if (is_subtotal or is_total) else FILL_FORECAST
                    c.border = border

        # Note
        if note:
            ws.cell(row=current_row, column=COL_NOTE, value=note).font = FONT_SMALL_NOTE

        this_row = current_row
        current_row += 1
        return this_row

    def empty_row():
        nonlocal current_row
        current_row += 1

    def margin_row(label, margin_data, forecast_formulas=None, note=""):
        """Lighter italic row for margins/growth rates."""
        nonlocal current_row

        c = ws.cell(row=current_row, column=COL_LABEL, value=label)
        c.font = Font(name="Arial", size=9, italic=True, color="666666")

        for year in YEARS_ACTUAL:
            col = year_col(year)
            val = margin_data.get(year)
            if val is not None:
                c = ws.cell(row=current_row, column=col, value=val)
                c.font = Font(name="Arial", size=9, italic=True, color="666666")
                c.number_format = FMT_PCT
                c.alignment = ALIGN_RIGHT

        if forecast_formulas:
            for year in YEARS_FORECAST:
                col = year_col(year)
                formula = forecast_formulas.get(year)
                if formula:
                    c = ws.cell(row=current_row, column=col, value=formula)
                    c.font = Font(name="Arial", size=9, italic=True, color="008000")
                    c.number_format = FMT_PCT
                    c.alignment = ALIGN_RIGHT

        if note:
            ws.cell(row=current_row, column=COL_NOTE, value=note).font = FONT_SMALL_NOTE

        current_row += 1

    # =========================================================================
    # BUILD THE INCOME STATEMENT
    # =========================================================================

    # --- REVENUE SECTION ---
    section_header("REVENUE")

    # Row references for formulas (tracked as we build)
    # Revenue row
    rev_row = data_row(
        "Total Revenue",
        actuals_key="Total Revenue",
        forecast_formulas={
            2026: f"=Assumptions!J{ASSUMP['revenue']}",
            2027: f"=Assumptions!K{ASSUMP['revenue']}",
            2028: f"=Assumptions!L{ASSUMP['revenue']}",
        },
        note="Source: Assumptions tab",
        is_subtotal=True,
    )

    # Revenue growth margin row
    rev_growth = {}
    for i, year in enumerate(YEARS_ACTUAL):
        if i > 0:
            rev_growth[year] = calc_growth(IS_DATA["Total Revenue"], year, YEARS_ACTUAL[i-1])
    margin_row(
        "  YoY Growth (%)",
        rev_growth,
        forecast_formulas={
            2026: f"=(J{rev_row}-I{rev_row})/I{rev_row}",
            2027: f"=(K{rev_row}-J{rev_row})/J{rev_row}",
            2028: f"=(L{rev_row}-K{rev_row})/K{rev_row}",
        },
    )

    empty_row()

    # --- COST OF REVENUE ---
    section_header("COST OF REVENUE")

    # COGS = Revenue × (1 - Gross Margin)
    cogs_row = data_row(
        "Cost of Revenue",
        actuals_key="Cost of Revenue",
        forecast_formulas={
            2026: f"=J{rev_row}*(1-Assumptions!J{ASSUMP['gross_margin']})",
            2027: f"=K{rev_row}*(1-Assumptions!K{ASSUMP['gross_margin']})",
            2028: f"=L{rev_row}*(1-Assumptions!L{ASSUMP['gross_margin']})",
        },
        note="Revenue × (1 - Gross Margin %)",
    )

    # COGS margin
    cogs_margin = {}
    for year in YEARS_ACTUAL:
        cogs_margin[year] = calc_margin(IS_DATA["Cost of Revenue"], IS_DATA["Total Revenue"], year)
    margin_row(
        "  % of Revenue",
        cogs_margin,
        forecast_formulas={
            2026: f"=J{cogs_row}/J{rev_row}",
            2027: f"=K{cogs_row}/K{rev_row}",
            2028: f"=L{cogs_row}/L{rev_row}",
        },
    )

    empty_row()

    # --- GROSS PROFIT ---
    gp_row = data_row(
        "Gross Profit",
        actuals_key="Gross Profit",
        forecast_formulas={
            2026: f"=J{rev_row}-J{cogs_row}",
            2027: f"=K{rev_row}-K{cogs_row}",
            2028: f"=L{rev_row}-L{cogs_row}",
        },
        is_subtotal=True,
        note="Revenue - COGS",
    )

    # Gross margin
    gm = {}
    for year in YEARS_ACTUAL:
        gm[year] = calc_margin(IS_DATA["Gross Profit"], IS_DATA["Total Revenue"], year)
    margin_row(
        "  Gross Margin (%)",
        gm,
        forecast_formulas={
            2026: f"=J{gp_row}/J{rev_row}",
            2027: f"=K{gp_row}/K{rev_row}",
            2028: f"=L{gp_row}/L{rev_row}",
        },
        note="Should match Assumptions tab gross margin",
    )

    empty_row()

    # --- OPERATING EXPENSES ---
    section_header("OPERATING EXPENSES")

    # Total OpEx = Revenue × OpEx % from Assumptions
    opex_row = data_row(
        "Total Operating Expenses",
        actuals_key="Operating Expenses",
        forecast_formulas={
            2026: f"=J{rev_row}*Assumptions!J{ASSUMP['total_opex_pct']}",
            2027: f"=K{rev_row}*Assumptions!K{ASSUMP['total_opex_pct']}",
            2028: f"=L{rev_row}*Assumptions!L{ASSUMP['total_opex_pct']}",
        },
        is_subtotal=True,
        note="Revenue × Total OpEx % (from Assumptions)",
    )

    # OpEx margin
    opex_margin = {}
    for year in YEARS_ACTUAL:
        opex_margin[year] = calc_margin(IS_DATA["Operating Expenses"], IS_DATA["Total Revenue"], year)
    margin_row(
        "  % of Revenue",
        opex_margin,
        forecast_formulas={
            2026: f"=J{opex_row}/J{rev_row}",
            2027: f"=K{opex_row}/K{rev_row}",
            2028: f"=L{opex_row}/L{rev_row}",
        },
    )

    empty_row()

    # --- OPERATING INCOME ---
    oi_row = data_row(
        "Operating Income (EBIT)",
        actuals_key="Operating Income",
        forecast_formulas={
            2026: f"=J{gp_row}-J{opex_row}",
            2027: f"=K{gp_row}-K{opex_row}",
            2028: f"=L{gp_row}-L{opex_row}",
        },
        is_subtotal=True,
        note="Gross Profit - Operating Expenses",
    )

    # Operating margin
    om = {}
    for year in YEARS_ACTUAL:
        om[year] = calc_margin(IS_DATA["Operating Income"], IS_DATA["Total Revenue"], year)
    margin_row(
        "  Operating Margin (%)",
        om,
        forecast_formulas={
            2026: f"=J{oi_row}/J{rev_row}",
            2027: f"=K{oi_row}/K{rev_row}",
            2028: f"=L{oi_row}/L{rev_row}",
        },
    )

    empty_row()

    # --- NON-OPERATING ITEMS ---
    section_header("NON-OPERATING ITEMS")

    int_inc_row = data_row(
        "Interest Income",
        actuals_key="Interest Income",
        forecast_formulas={
            2026: "=380",  # Lower rates per mgmt guidance
            2027: "=350",
            2028: "=330",
        },
        note="Lower rates headwind per Q4'25 guidance",
    )

    int_exp_row = data_row(
        "Interest Expense",
        actuals_key="Interest Expense",
        forecast_formulas={
            2026: f"=-Assumptions!J{ASSUMP['interest_expense']}",
            2027: f"=-Assumptions!K{ASSUMP['interest_expense']}",
            2028: f"=-Assumptions!L{ASSUMP['interest_expense']}",
        },
        note="From Assumptions tab (shown as negative)",
    )

    other_row = data_row(
        "Other Income (Expense)",
        actuals_key="Other Income (Expense)",
        forecast_formulas={
            2026: "=0",   # Conservative: assume zero
            2027: "=0",
            2028: "=0",
        },
        note="Volatile & unpredictable; conservatively set to $0",
    )

    empty_row()

    # --- INCOME BEFORE TAXES ---
    ebt_row = data_row(
        "Income Before Taxes",
        actuals_key="Income Before Taxes",
        forecast_formulas={
            2026: f"=J{oi_row}+J{int_inc_row}+J{int_exp_row}+J{other_row}",
            2027: f"=K{oi_row}+K{int_inc_row}+K{int_exp_row}+K{other_row}",
            2028: f"=L{oi_row}+L{int_inc_row}+L{int_exp_row}+L{other_row}",
        },
        is_subtotal=True,
        note="EBIT + Interest Inc + Interest Exp + Other",
    )

    empty_row()

    # --- TAXES ---
    section_header("TAXES")

    # Tax = EBT × Tax Rate
    tax_row = data_row(
        "Income Tax Expense",
        actuals_key="Income Tax Expense",
        forecast_formulas={
            2026: f"=J{ebt_row}*Assumptions!J{ASSUMP['tax_rate']}",
            2027: f"=K{ebt_row}*Assumptions!K{ASSUMP['tax_rate']}",
            2028: f"=L{ebt_row}*Assumptions!L{ASSUMP['tax_rate']}",
        },
        note="EBT × Effective Tax Rate (from Assumptions)",
    )

    # Effective tax rate
    eff_tax = {}
    for year in YEARS_ACTUAL:
        ebt = IS_DATA["Income Before Taxes"].get(year, 0)
        tax = IS_DATA["Income Tax Expense"].get(year, 0)
        if ebt != 0:
            eff_tax[year] = tax / ebt
    margin_row(
        "  Effective Tax Rate (%)",
        eff_tax,
        forecast_formulas={
            2026: f"=IF(J{ebt_row}=0,0,J{tax_row}/J{ebt_row})",
            2027: f"=IF(K{ebt_row}=0,0,K{tax_row}/K{ebt_row})",
            2028: f"=IF(L{ebt_row}=0,0,L{tax_row}/L{ebt_row})",
        },
    )

    empty_row()

    # --- NET INCOME ---
    ni_row = data_row(
        "Net Income",
        actuals_key="Net Income",
        forecast_formulas={
            2026: f"=J{ebt_row}-J{tax_row}",
            2027: f"=K{ebt_row}-K{tax_row}",
            2028: f"=L{ebt_row}-L{tax_row}",
        },
        is_total=True,
        note="EBT - Tax Expense",
    )

    # Net margin
    nm = {}
    for year in YEARS_ACTUAL:
        nm[year] = calc_margin(IS_DATA["Net Income"], IS_DATA["Total Revenue"], year)
    margin_row(
        "  Net Margin (%)",
        nm,
        forecast_formulas={
            2026: f"=J{ni_row}/J{rev_row}",
            2027: f"=K{ni_row}/K{rev_row}",
            2028: f"=L{ni_row}/L{rev_row}",
        },
    )

    # Net income growth
    ni_growth = {}
    for i, year in enumerate(YEARS_ACTUAL):
        if i > 0:
            ni_growth[year] = calc_growth(IS_DATA["Net Income"], year, YEARS_ACTUAL[i-1])
    margin_row(
        "  YoY Growth (%)",
        ni_growth,
        forecast_formulas={
            2026: f"=(J{ni_row}-I{ni_row})/ABS(I{ni_row})",
            2027: f"=(K{ni_row}-J{ni_row})/ABS(J{ni_row})",
            2028: f"=(L{ni_row}-K{ni_row})/ABS(K{ni_row})",
        },
    )

    empty_row()

    # --- PER SHARE DATA ---
    section_header("PER SHARE DATA")

    shares_row = data_row(
        "Diluted Shares Outstanding (M)",
        actuals_key="Diluted Shares (M)",
        forecast_formulas={
            # At ~$40-50/share, $6B buys ~130M shares; SBC adds ~25M; net ~105M reduction/yr
            2026: "=830",
            2027: "=725",
            2028: "=625",
        },
        fmt=FMT_SHARES,
        note="Accelerated buyback at lower share price",
    )

    eps_row = data_row(
        "Diluted EPS",
        actuals_key="Diluted EPS",
        forecast_formulas={
            2026: f"=J{ni_row}/J{shares_row}",
            2027: f"=K{ni_row}/K{shares_row}",
            2028: f"=L{ni_row}/L{shares_row}",
        },
        fmt=FMT_PER_SHARE,
        is_subtotal=True,
        note="Net Income / Diluted Shares",
    )

    # EPS growth
    eps_growth = {}
    for i, year in enumerate(YEARS_ACTUAL):
        if i > 0:
            eps_growth[year] = calc_growth(IS_DATA["Diluted EPS"], year, YEARS_ACTUAL[i-1])
    margin_row(
        "  EPS Growth (%)",
        eps_growth,
        forecast_formulas={
            2026: f"=(J{eps_row}-I{eps_row})/ABS(I{eps_row})",
            2027: f"=(K{eps_row}-J{eps_row})/ABS(J{eps_row})",
            2028: f"=(L{eps_row}-K{eps_row})/ABS(K{eps_row})",
        },
    )

    # =========================================================================
    # PRINT ROW MAP (for next tabs to reference)
    # =========================================================================
    empty_row()
    empty_row()
    r = current_row
    ws.cell(row=r, column=2, value="Row Reference Map (for model linking):").font = FONT_SMALL_NOTE
    r += 1
    refs = [
        f"Revenue: Row {rev_row}",
        f"COGS: Row {cogs_row}",
        f"Gross Profit: Row {gp_row}",
        f"OpEx: Row {opex_row}",
        f"Operating Income: Row {oi_row}",
        f"EBT: Row {ebt_row}",
        f"Tax: Row {tax_row}",
        f"Net Income: Row {ni_row}",
        f"Shares: Row {shares_row}",
        f"EPS: Row {eps_row}",
    ]
    for ref in refs:
        ws.cell(row=r, column=2, value=ref).font = Font(name="Arial", size=8, color="BBBBBB")
        r += 1

    # =========================================================================
    # FREEZE PANES
    # =========================================================================
    ws.freeze_panes = "C5"

    print(f"\n  Income Statement Row Map:")
    print(f"    Revenue:          Row {rev_row}")
    print(f"    COGS:             Row {cogs_row}")
    print(f"    Gross Profit:     Row {gp_row}")
    print(f"    OpEx:             Row {opex_row}")
    print(f"    Operating Income: Row {oi_row}")
    print(f"    Interest Income:  Row {int_inc_row}")
    print(f"    Interest Expense: Row {int_exp_row}")
    print(f"    Other Inc/Exp:    Row {other_row}")
    print(f"    EBT:              Row {ebt_row}")
    print(f"    Tax:              Row {tax_row}")
    print(f"    Net Income:       Row {ni_row}")
    print(f"    Shares:           Row {shares_row}")
    print(f"    EPS:              Row {eps_row}")

    return ws


# =============================================================================
# MAIN
# =============================================================================
def main():
    print(f"Opening model: {MODEL_PATH}")

    if not os.path.exists(MODEL_PATH):
        print(f"  ✗ File not found! Run 03_build_excel_model.py first.")
        return

    wb = load_workbook(MODEL_PATH)
    print(f"  ✓ Loaded. Existing tabs: {', '.join(wb.sheetnames)}")

    build_income_statement(wb)

    wb.save(MODEL_PATH)
    print(f"\n  ✓ Income Statement tab added and saved.")
    print(f"  ✓ File: {MODEL_PATH}")
    print(f"\n  → Open in Excel. Forecast columns (J-L) contain formulas")
    print(f"    linking to the Assumptions tab (green font).")
    print(f"  → Recalculate in Excel: Ctrl+Shift+F9")
    print(f"\n  Next step: Balance Sheet tab")


if __name__ == "__main__":
    main()
