"""
PayPal (PYPL) - Excel Financial Model Builder
==============================================
Phase 3, Step 1: Cover + Assumptions tabs

Industry-standard formatting:
  - Blue font: hardcoded inputs (editable assumptions)
  - Black font: formulas/calculations
  - Green font: cross-sheet links
  - Yellow background: key assumptions needing attention
  - Years as text, currency as $#,##0, percentages as 0.0%

Run from scripts/ folder.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

# =============================================================================
# CONFIGURATION
# =============================================================================
BASE_DIR = os.path.dirname(os.path.dirname(__file__))
OUTPUT_PATH = os.path.join(BASE_DIR, "model", "PYPL_Financial_Model.xlsx")
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
# Fonts
FONT_TITLE = Font(name="Arial", size=20, bold=True, color="000000")
FONT_SUBTITLE = Font(name="Arial", size=14, color="444444")
FONT_SECTION = Font(name="Arial", size=11, bold=True, color="000000")
FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_YEAR_ACTUAL = Font(name="Arial", size=10, bold=True, color="000000")
FONT_YEAR_FORECAST = Font(name="Arial", size=10, bold=True, color="0000FF")
FONT_LABEL = Font(name="Arial", size=10, color="000000")
FONT_LABEL_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
FONT_INPUT_BLUE = Font(name="Arial", size=10, color="0000FF")  # Hardcoded inputs
FONT_FORMULA = Font(name="Arial", size=10, color="000000")  # Formulas
FONT_GREEN_LINK = Font(name="Arial", size=10, color="008000")  # Cross-sheet links
FONT_COVER_INFO = Font(name="Arial", size=12, color="000000")
FONT_COVER_LABEL = Font(name="Arial", size=12, bold=True, color="444444")
FONT_SMALL_NOTE = Font(name="Arial", size=9, italic=True, color="888888")

# Fills
FILL_HEADER = PatternFill("solid", fgColor="2E4057")  # Dark navy
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")  # Light blue
FILL_INPUT = PatternFill("solid", fgColor="FFFFCC")  # Light yellow for key inputs
FILL_ACTUAL = PatternFill("solid", fgColor="F2F2F2")  # Light gray for actuals
FILL_FORECAST = PatternFill("solid", fgColor="E8F0FE")  # Light blue for forecast
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_COVER_ACCENT = PatternFill("solid", fgColor="2E4057")
FILL_SUBTOTAL = PatternFill("solid", fgColor="E8E8E8")

# Alignment
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Borders
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color="000000"),
)
SECTION_BORDER = Border(
    top=Side(style="medium", color="2E4057"),
    bottom=Side(style="thin", color="2E4057"),
)

# Number formats
FMT_CURRENCY = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_PCT_INPUT = '0.0%'
FMT_MULTIPLE = '0.0"x"'
FMT_NUMBER = '#,##0;(#,##0);"-"'
FMT_YEAR = '@'  # Text format for years


# =============================================================================
# HISTORICAL DATA (from extraction — all values in USD Millions)
# Source: yfinance + SEC EDGAR 10-K filings
# =============================================================================
HISTORICAL = {
    # INCOME STATEMENT
    "revenue": {2019: 17772, 2020: 21454, 2021: 25371, 2022: 27518, 2023: 29771, 2024: 31797, 2025: 33172},
    "cogs": {2019: 10004, 2020: 12253, 2021: 14233, 2022: 15326, 2023: 16395, 2024: 17141, 2025: 17862},
    "gross_profit": {2019: 7768, 2020: 9201, 2021: 11138, 2022: 12192, 2023: 13376, 2024: 14656, 2025: 15310},
    "operating_income": {2019: 2667, 2020: 3289, 2021: 4202, 2022: 4136, 2023: 5028, 2024: 5718, 2025: 6008},
    "net_income": {2019: 2459, 2020: 4202, 2021: 4169, 2022: 2419, 2023: 4246, 2024: 4573, 2025: 4310},
    "interest_expense": {2019: 234, 2020: 264, 2021: 232, 2022: 309, 2023: 365, 2024: 392, 2025: 398},

    # BALANCE SHEET
    "total_assets": {2019: 51333, 2020: 70379, 2021: 75803, 2022: 78624, 2023: 82166, 2024: 78725, 2025: 80173},
    "total_equity": {2019: 16885, 2020: 20019, 2021: 21727, 2022: 20274, 2023: 21051, 2024: 20417, 2025: 20256},
    "total_current_assets": {2019: 38495, 2020: 50995, 2021: 18029, 2022: 57424, 2023: 62569, 2024: 58206, 2025: 59759},
    "total_current_liabilities": {2019: 26919, 2020: 38447, 2021: 43029, 2022: 45008, 2023: 48466, 2024: 45490, 2025: 46443},
    "cash": {2019: 7349, 2020: 4794, 2021: 5197, 2022: 7776, 2023: 9081, 2024: 6662, 2025: 8049},
    "long_term_debt": {2019: 4965, 2020: 8939, 2021: 8049, 2022: 10417, 2023: 9676, 2024: 9879, 2025: 9987},
    "short_term_debt": {2019: 0, 2020: 0, 2021: 0, 2022: 0, 2023: 0, 2024: 0, 2025: 0},
    "goodwill": {2019: 6212, 2020: 9135, 2021: 11454, 2022: 11209, 2023: 11026, 2024: 10837, 2025: 10864},
    "ppe_net": {2019: 1693, 2020: 1807, 2021: 1909, 2022: 2201, 2023: 1882, 2024: 1581, 2025: 1909},

    # CASH FLOW
    "cfo": {2019: 4562, 2020: 5853, 2021: 6343, 2022: 5813, 2023: 4843, 2024: 6400, 2025: 7120},
    "capex": {2019: 704, 2020: 866, 2021: 908, 2022: 706, 2023: 596, 2024: 516, 2025: 505},
    "da": {2019: 810, 2020: 1011, 2021: 1200, 2022: 1274, 2023: 1290, 2024: 1259, 2025: 1280},
    "sbc": {2019: 1021, 2020: 1378, 2021: 1510, 2022: 1456, 2023: 1390, 2024: 1399, 2025: 1350},
    "buybacks": {2019: 3333, 2020: 1021, 2021: 3397, 2022: 4200, 2023: 5100, 2024: 5400, 2025: 6000},
}

YEARS_ACTUAL = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
YEARS_FORECAST = [2026, 2027, 2028]
ALL_YEARS = YEARS_ACTUAL + YEARS_FORECAST


def safe_div(a, b):
    if b is None or b == 0 or a is None:
        return None
    return a / b


# =============================================================================
# BUILD COVER TAB
# =============================================================================
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = "2E4057"

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 5

    # White background
    for row in range(1, 35):
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = FILL_WHITE

    # Company name bar
    ws.merge_cells("A3:D3")
    ws["A3"].fill = FILL_COVER_ACCENT
    ws.row_dimensions[3].height = 50
    ws.merge_cells("A4:D4")
    ws["A4"].fill = FILL_COVER_ACCENT
    ws.row_dimensions[4].height = 8

    # Title
    ws.merge_cells("B6:C6")
    c = ws["B6"]
    c.value = "PayPal Holdings, Inc. (PYPL)"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT

    ws.merge_cells("B7:C7")
    c = ws["B7"]
    c.value = "3-Statement Financial Model & Investment Analysis"
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_LEFT

    ws.merge_cells("B8:C8")
    c = ws["B8"]
    c.value = "Equity Research — Investment Recommendation"
    c.font = Font(name="Arial", size=11, italic=True, color="666666")
    c.alignment = ALIGN_LEFT

    # Separator
    ws.row_dimensions[9].height = 5

    # Info block
    info_rows = [
        ("Sector:", "Financial Technology / Digital Payments"),
        ("Exchange:", "NASDAQ: PYPL"),
        ("Fiscal Year End:", "December 31"),
        ("Currency:", "USD (values in millions unless noted)"),
        ("", ""),
        ("Analyst:", "Carlos — Investment Analysis Portfolio Project"),
        ("Date:", "February 2026"),
        ("Data Sources:", "SEC EDGAR 10-K Filings, yfinance, Company Reports"),
        ("", ""),
        ("Historical Period:", "FY2019 — FY2025 (7 years)"),
        ("Forecast Period:", "FY2026E — FY2028E (3 years)"),
        ("Scenarios:", "Base Case / Bull Case / Bear Case"),
        ("Valuation Method:", "Discounted Cash Flow (DCF) — FCFF"),
    ]

    start_row = 11
    for i, (label, value) in enumerate(info_rows):
        r = start_row + i
        ws.cell(row=r, column=2, value=label).font = FONT_COVER_LABEL
        ws.cell(row=r, column=3, value=value).font = FONT_COVER_INFO
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.cell(row=r, column=3).alignment = ALIGN_LEFT

    # Recommendation placeholder
    r = start_row + len(info_rows) + 1
    ws.merge_cells(f"B{r}:C{r}")
    ws.row_dimensions[r].height = 6
    ws.cell(row=r, column=2).border = Border(bottom=Side(style="medium", color="2E4057"))
    ws.cell(row=r, column=3).border = Border(bottom=Side(style="medium", color="2E4057"))

    r += 2
    ws.cell(row=r, column=2, value="RECOMMENDATION").font = Font(name="Arial", size=14, bold=True, color="2E4057")
    r += 1
    ws.cell(row=r, column=2, value="To be determined after model completion").font = Font(name="Arial", size=12, italic=True, color="999999")
    r += 1
    ws.cell(row=r, column=2, value="Target Price: $___  |  Current Price: $___  |  Upside: ___%").font = Font(name="Arial", size=11, color="999999")

    # Tab guide
    r += 3
    ws.cell(row=r, column=2, value="MODEL NAVIGATION").font = Font(name="Arial", size=11, bold=True, color="2E4057")
    ws.cell(row=r, column=2).border = Border(bottom=Side(style="thin", color="2E4057"))
    ws.cell(row=r, column=3).border = Border(bottom=Side(style="thin", color="2E4057"))

    tabs = [
        ("Cover", "This page — project overview"),
        ("Assumptions", "All forecast drivers and inputs (start here)"),
        ("Income Statement", "Historical + projected P&L"),
        ("Balance Sheet", "Historical + projected BS"),
        ("Cash Flow", "Historical + projected CF"),
        ("Ratios", "Financial ratio analysis"),
        ("DCF", "Discounted Cash Flow valuation"),
        ("Scenarios", "Bull / Base / Bear comparison"),
        ("Investment Memo", "Executive summary & recommendation"),
    ]
    for i, (tab, desc) in enumerate(tabs):
        ws.cell(row=r + 1 + i, column=2, value=tab).font = Font(name="Arial", size=10, bold=True, color="000000")
        ws.cell(row=r + 1 + i, column=3, value=desc).font = Font(name="Arial", size=10, color="666666")

    # Formatting legend
    r = r + len(tabs) + 2
    ws.cell(row=r, column=2, value="FORMATTING LEGEND").font = Font(name="Arial", size=11, bold=True, color="2E4057")
    ws.cell(row=r, column=2).border = Border(bottom=Side(style="thin", color="2E4057"))
    ws.cell(row=r, column=3).border = Border(bottom=Side(style="thin", color="2E4057"))

    legends = [
        ("Blue text", "Hardcoded input — editable assumption", "0000FF", None),
        ("Black text", "Formula / calculation", "000000", None),
        ("Green text", "Link to another worksheet", "008000", None),
        ("Yellow background", "Key assumption needing review", "000000", "FFFFCC"),
    ]
    for i, (label, desc, color, bg) in enumerate(legends):
        c = ws.cell(row=r + 1 + i, column=2, value=label)
        c.font = Font(name="Arial", size=10, bold=True, color=color)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=r + 1 + i, column=3, value=desc).font = Font(name="Arial", size=10, color="666666")

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.print_area = "A1:D50"

    return ws


# =============================================================================
# BUILD ASSUMPTIONS TAB
# =============================================================================
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = "0000FF"

    # Column layout:
    # A: empty spacer (width 2)
    # B: Label (width 35)
    # C-I: FY2019-FY2025 actuals (width 14 each)
    # J-L: FY2026E-FY2028E forecasts (width 14 each)
    # M: Notes (width 30)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    for i, col_letter in enumerate(["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]):
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["M"].width = 35

    # Helper: column index for year
    def year_col(year):
        if year in YEARS_ACTUAL:
            return 3 + YEARS_ACTUAL.index(year)  # C=3 through I=9
        else:
            return 10 + YEARS_FORECAST.index(year)  # J=10 through L=12

    def note_col():
        return 13  # M

    # =========================================================================
    # TITLE ROW
    # =========================================================================
    row = 1
    ws.merge_cells(f"B{row}:L{row}")
    c = ws.cell(row=row, column=2, value="PayPal (PYPL) — Model Assumptions")
    c.font = Font(name="Arial", size=14, bold=True, color="2E4057")
    c.alignment = ALIGN_LEFT

    row = 2
    ws.cell(row=row, column=2, value="All monetary values in USD Millions ($M)").font = FONT_SMALL_NOTE
    ws.cell(row=row, column=note_col(), value="Notes / Source").font = Font(name="Arial", size=10, bold=True, italic=True, color="888888")

    # =========================================================================
    # YEAR HEADERS (Row 3)
    # =========================================================================
    row = 3
    ws.cell(row=row, column=2, value="").fill = FILL_HEADER
    ws.cell(row=row, column=note_col(), value="").fill = FILL_HEADER

    for year in YEARS_ACTUAL:
        col = year_col(year)
        c = ws.cell(row=row, column=col, value=str(year))
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.number_format = FMT_YEAR

    for year in YEARS_FORECAST:
        col = year_col(year)
        c = ws.cell(row=row, column=col, value=f"{year}E")
        c.font = Font(name="Arial", size=10, bold=True, color="CCDDFF")
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

    # Period type labels
    row = 4
    ws.cell(row=row, column=2, value="").font = FONT_LABEL
    for year in YEARS_ACTUAL:
        c = ws.cell(row=row, column=year_col(year), value="Actual")
        c.font = Font(name="Arial", size=8, italic=True, color="888888")
        c.alignment = ALIGN_CENTER
    for year in YEARS_FORECAST:
        c = ws.cell(row=row, column=year_col(year), value="Forecast")
        c.font = Font(name="Arial", size=8, italic=True, color="0000FF")
        c.alignment = ALIGN_CENTER

    # =========================================================================
    # HELPER FUNCTIONS
    # =========================================================================
    def add_section_header(row, title):
        ws.merge_cells(f"B{row}:L{row}")
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(name="Arial", size=11, bold=True, color="2E4057")
        c.fill = FILL_SECTION
        c.alignment = ALIGN_LEFT
        c.border = SECTION_BORDER
        for col in range(3, 14):
            ws.cell(row=row, column=col).fill = FILL_SECTION
            ws.cell(row=row, column=col).border = SECTION_BORDER
        return row + 1

    def add_data_row(row, label, data_key=None, values=None, fmt="currency",
                     forecast_values=None, note="", is_calculated=False, is_subtotal=False):
        """
        Add a row with historical actuals and forecast inputs.
        - data_key: key in HISTORICAL dict for actuals
        - values: pre-calculated dict {year: value} (alternative to data_key)
        - forecast_values: dict {year: value} for forecast period inputs
        - fmt: 'currency', 'pct', 'multiple', 'number'
        - is_calculated: if True, actuals are formula-derived (black font)
        - is_subtotal: bold + border
        """
        # Label
        c = ws.cell(row=row, column=2, value=label)
        c.font = FONT_LABEL_BOLD if is_subtotal else FONT_LABEL
        c.alignment = ALIGN_LEFT
        if is_subtotal:
            c.border = BOTTOM_BORDER

        # Number format
        if fmt == "currency":
            nf = FMT_CURRENCY
        elif fmt == "pct":
            nf = FMT_PCT
        elif fmt == "multiple":
            nf = FMT_MULTIPLE
        else:
            nf = FMT_NUMBER

        # Actual values
        actual_data = values if values else (HISTORICAL.get(data_key, {}) if data_key else {})
        for year in YEARS_ACTUAL:
            col = year_col(year)
            val = actual_data.get(year)
            if val is not None:
                c = ws.cell(row=row, column=col, value=val)
                c.font = FONT_FORMULA if is_calculated else FONT_INPUT_BLUE
                c.number_format = nf
                c.alignment = ALIGN_RIGHT
                c.fill = FILL_ACTUAL
                if is_subtotal:
                    c.font = Font(name="Arial", size=10, bold=True, color="000000")
                    c.border = BOTTOM_BORDER

        # Forecast values (always blue + yellow background for key inputs)
        if forecast_values:
            for year in YEARS_FORECAST:
                col = year_col(year)
                val = forecast_values.get(year)
                if val is not None:
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = FONT_INPUT_BLUE
                    c.number_format = nf
                    c.alignment = ALIGN_RIGHT
                    c.fill = FILL_INPUT
                    if is_subtotal:
                        c.font = Font(name="Arial", size=10, bold=True, color="0000FF")
                        c.border = BOTTOM_BORDER

        # Note
        if note:
            c = ws.cell(row=row, column=note_col(), value=note)
            c.font = FONT_SMALL_NOTE
            c.alignment = ALIGN_LEFT

        return row + 1

    def add_empty_row(row):
        return row + 1

    # =========================================================================
    # SECTION 1: REVENUE DRIVERS
    # =========================================================================
    row = 6
    row = add_section_header(row, "REVENUE DRIVERS")

    row = add_data_row(row, "Total Revenue ($M)", data_key="revenue",
                       forecast_values={2026: 34170, 2027: 35540, 2028: 37320},
                       note="Post-Q4'25: ~3% '26, ~4% '27, ~5% '28 recovery",
                       is_subtotal=True)

    # Revenue growth (calculated from revenue)
    rev = HISTORICAL["revenue"]
    rev_growth = {}
    for i, year in enumerate(YEARS_ACTUAL):
        if i > 0:
            prev = YEARS_ACTUAL[i - 1]
            rev_growth[year] = (rev[year] - rev[prev]) / rev[prev]

    row = add_data_row(row, "  Revenue Growth (%)", values=rev_growth, fmt="pct",
                       forecast_values={2026: 0.030, 2027: 0.040, 2028: 0.050},
                       is_calculated=True,
                       note="Weak '26 guidance; gradual recovery assumed")

    # Gross margin
    gm = {}
    for year in YEARS_ACTUAL:
        gp = HISTORICAL["gross_profit"].get(year)
        r = HISTORICAL["revenue"].get(year)
        if gp and r:
            gm[year] = gp / r

    row = add_data_row(row, "  Gross Margin (%)", values=gm, fmt="pct",
                       forecast_values={2026: 0.460, 2027: 0.462, 2028: 0.465},
                       is_calculated=True,
                       note="Flat near-term; modest expansion from mix shift")

    row = add_empty_row(row)

    # =========================================================================
    # SECTION 2: OPERATING EXPENSE DRIVERS
    # =========================================================================
    row = add_section_header(row, "OPERATING EXPENSE DRIVERS (as % of Revenue)")

    # Calculate historical OpEx ratios
    # Total OpEx implied = Gross Profit - Operating Income
    total_opex_pct = {}
    for year in YEARS_ACTUAL:
        gp = HISTORICAL["gross_profit"].get(year)
        oi = HISTORICAL["operating_income"].get(year)
        r = HISTORICAL["revenue"].get(year)
        if gp and oi and r:
            total_opex_pct[year] = (gp - oi) / r

    row = add_data_row(row, "Total OpEx (% of Revenue)", values=total_opex_pct, fmt="pct",
                       forecast_values={2026: 0.280, 2027: 0.275, 2028: 0.270},
                       is_calculated=True, is_subtotal=True,
                       note="CEO transition costs; limited leverage near-term")

    # Operating margin
    op_margin = {}
    for year in YEARS_ACTUAL:
        oi = HISTORICAL["operating_income"].get(year)
        r = HISTORICAL["revenue"].get(year)
        if oi and r:
            op_margin[year] = oi / r

    row = add_data_row(row, "  Operating Margin (%)", values=op_margin, fmt="pct",
                       forecast_values={2026: 0.180, 2027: 0.187, 2028: 0.195},
                       is_calculated=True,
                       note="Gross Margin - Total OpEx %; slow recovery")

    # D&A
    da_pct = {}
    for year in YEARS_ACTUAL:
        d = HISTORICAL["da"].get(year)
        r = HISTORICAL["revenue"].get(year)
        if d and r:
            da_pct[year] = d / r

    row = add_data_row(row, "  D&A (% of Revenue)", values=da_pct, fmt="pct",
                       forecast_values={2026: 0.038, 2027: 0.037, 2028: 0.036},
                       is_calculated=True,
                       note="Stable; slight decline")

    # SBC
    sbc_pct = {}
    for year in YEARS_ACTUAL:
        s = HISTORICAL["sbc"].get(year)
        r = HISTORICAL["revenue"].get(year)
        if s and r:
            sbc_pct[year] = s / r

    row = add_data_row(row, "  SBC (% of Revenue)", values=sbc_pct, fmt="pct",
                       forecast_values={2026: 0.040, 2027: 0.039, 2028: 0.038},
                       is_calculated=True,
                       note="Elevated during CEO transition")

    row = add_empty_row(row)

    # =========================================================================
    # SECTION 3: TAX & INTEREST
    # =========================================================================
    row = add_section_header(row, "TAX & INTEREST")

    # Effective tax rate
    tax_rate = {}
    for year in YEARS_ACTUAL:
        ni = HISTORICAL["net_income"].get(year)
        oi = HISTORICAL["operating_income"].get(year)
        ie = HISTORICAL["interest_expense"].get(year)
        if ni and oi:
            ebt = oi - (ie or 0)
            if ebt != 0:
                tax = ebt - ni
                tax_rate[year] = tax / ebt

    row = add_data_row(row, "Effective Tax Rate (%)", values=tax_rate, fmt="pct",
                       forecast_values={2026: 0.175, 2027: 0.175, 2028: 0.175},
                       is_calculated=True,
                       note="Assumed stable effective rate")

    row = add_data_row(row, "Interest Expense ($M)", data_key="interest_expense",
                       forecast_values={2026: 410, 2027: 420, 2028: 420},
                       note="Slightly higher; debt remains elevated")

    row = add_empty_row(row)

    # =========================================================================
    # SECTION 4: BALANCE SHEET DRIVERS
    # =========================================================================
    row = add_section_header(row, "BALANCE SHEET DRIVERS")

    # CapEx
    row = add_data_row(row, "Capital Expenditures ($M)", data_key="capex",
                       forecast_values={2026: 900, 2027: 850, 2028: 800},
                       note="Mgmt guided ~$1B '26; investment cycle")

    capex_pct = {}
    for year in YEARS_ACTUAL:
        cx = HISTORICAL["capex"].get(year)
        r = HISTORICAL["revenue"].get(year)
        if cx and r:
            capex_pct[year] = cx / r

    row = add_data_row(row, "  CapEx (% of Revenue)", values=capex_pct, fmt="pct",
                       forecast_values={2026: 0.026, 2027: 0.024, 2028: 0.021},
                       is_calculated=True,
                       note="Higher investment cycle; normalizes by '28")

    # D&A absolute
    row = add_data_row(row, "Depreciation & Amortization ($M)", data_key="da",
                       forecast_values={2026: 1300, 2027: 1320, 2028: 1340},
                       note="Source: CF statement; stable")

    # Net PP&E
    row = add_data_row(row, "Net PP&E ($M)", data_key="ppe_net",
                       forecast_values={2026: None, 2027: None, 2028: None},
                       note="= Prior PP&E + CapEx - Depreciation (formula in BS)")

    row = add_empty_row(row)

    # Working capital items
    row = add_section_header(row, "WORKING CAPITAL DRIVERS")

    # Current ratio
    cr = {}
    for year in YEARS_ACTUAL:
        ca = HISTORICAL["total_current_assets"].get(year)
        cl = HISTORICAL["total_current_liabilities"].get(year)
        if ca and cl and cl != 0:
            cr[year] = ca / cl

    row = add_data_row(row, "Current Ratio", values=cr, fmt="multiple",
                       forecast_values={2026: None, 2027: None, 2028: None},
                       is_calculated=True,
                       note="= Current Assets / Current Liabilities (calc in BS)")

    row = add_empty_row(row)

    # =========================================================================
    # SECTION 5: DEBT & CAPITAL STRUCTURE
    # =========================================================================
    row = add_section_header(row, "DEBT & CAPITAL STRUCTURE")

    row = add_data_row(row, "Long-Term Debt ($M)", data_key="long_term_debt",
                       forecast_values={2026: 9900, 2027: 9800, 2028: 9700},
                       note="Minimal deleveraging; focus on buybacks")

    # Debt to equity
    dte = {}
    for year in YEARS_ACTUAL:
        ltd = HISTORICAL["long_term_debt"].get(year, 0)
        std = HISTORICAL["short_term_debt"].get(year, 0)
        eq = HISTORICAL["total_equity"].get(year)
        if eq and eq != 0:
            dte[year] = (ltd + std) / eq

    row = add_data_row(row, "  Debt / Equity", values=dte, fmt="multiple",
                       is_calculated=True,
                       note="Total debt / stockholders equity")

    # Net debt
    net_debt = {}
    for year in YEARS_ACTUAL:
        ltd = HISTORICAL["long_term_debt"].get(year, 0)
        std = HISTORICAL["short_term_debt"].get(year, 0)
        c = HISTORICAL["cash"].get(year, 0)
        net_debt[year] = ltd + std - c

    row = add_data_row(row, "  Net Debt ($M)", values=net_debt,
                       is_calculated=True,
                       note="Total Debt - Cash")

    # Share repurchases
    row = add_data_row(row, "Share Repurchases ($M)", data_key="buybacks",
                       forecast_values={2026: 6000, 2027: 6000, 2028: 6000},
                       note="Mgmt guided ~$6B/yr; maintained")

    row = add_empty_row(row)

    # =========================================================================
    # SECTION 6: DCF VALUATION INPUTS
    # =========================================================================
    row = add_section_header(row, "DCF VALUATION INPUTS")

    # These are single-value inputs, not time series
    dcf_inputs = [
        ("Risk-Free Rate (%)", 0.042, "10Y US Treasury yield"),
        ("Equity Risk Premium (%)", 0.055, "Damodaran 2025 estimate"),
        ("Beta (levered)", 1.40, "Elevated: CEO change, competitive pressure"),
        ("Cost of Equity (CAPM) (%)", None, "= Rf + Beta × ERP (formula)"),
        ("Pre-Tax Cost of Debt (%)", 0.042, "= Interest Expense / Avg Debt"),
        ("Tax Rate for WACC (%)", 0.175, "Same as effective tax rate"),
        ("Target Debt Weight (%)", 0.30, "Based on current capital structure"),
        ("Target Equity Weight (%)", 0.70, "= 1 - Debt Weight"),
        ("WACC (%)", None, "= E/V × Ke + D/V × Kd × (1-t) (formula)"),
        ("", None, ""),
        ("Terminal Growth Rate (%)", 0.015, "Conservative: below GDP, competitive risk"),
        ("Exit EV/EBITDA Multiple", None, "Alternative terminal value method"),
    ]

    for label, val, note in dcf_inputs:
        if label == "":
            row = add_empty_row(row)
            continue

        c = ws.cell(row=row, column=2, value=label)
        c.font = FONT_LABEL
        c.alignment = ALIGN_LEFT

        # Put value in column C (first data column)
        if val is not None:
            c = ws.cell(row=row, column=3, value=val)
            c.font = FONT_INPUT_BLUE
            c.fill = FILL_INPUT
            if "Rate" in label or "Premium" in label or "Weight" in label or "WACC" in label or "Growth" in label or "Cost" in label:
                c.number_format = FMT_PCT
            elif "Beta" in label:
                c.number_format = "0.00"
            elif "Multiple" in label:
                c.number_format = FMT_MULTIPLE
            c.alignment = ALIGN_RIGHT
        else:
            c = ws.cell(row=row, column=3, value="← formula")
            c.font = Font(name="Arial", size=9, italic=True, color="999999")
            c.alignment = ALIGN_CENTER

        # Note
        ws.cell(row=row, column=note_col(), value=note).font = FONT_SMALL_NOTE

        row += 1

    row = add_empty_row(row)

    # =========================================================================
    # SECTION 7: KEY CALCULATED METRICS (reference)
    # =========================================================================
    row = add_section_header(row, "KEY CALCULATED METRICS (for reference — formulas in other sheets)")

    # Free Cash Flow
    fcf = {}
    for year in YEARS_ACTUAL:
        cfo_val = HISTORICAL["cfo"].get(year)
        capex_val = HISTORICAL["capex"].get(year)
        if cfo_val and capex_val:
            fcf[year] = cfo_val - capex_val

    row = add_data_row(row, "Free Cash Flow ($M)", values=fcf,
                       is_calculated=True, is_subtotal=True,
                       note="= Cash from Operations - CapEx")

    # FCF margin
    fcf_margin = {}
    for year in YEARS_ACTUAL:
        f = fcf.get(year)
        r = HISTORICAL["revenue"].get(year)
        if f and r:
            fcf_margin[year] = f / r

    row = add_data_row(row, "  FCF Margin (%)", values=fcf_margin, fmt="pct",
                       is_calculated=True,
                       note="Strong cash generation — key PayPal strength")

    # EBITDA
    ebitda = {}
    for year in YEARS_ACTUAL:
        oi = HISTORICAL["operating_income"].get(year)
        da = HISTORICAL["da"].get(year)
        if oi and da:
            ebitda[year] = oi + da

    row = add_data_row(row, "EBITDA ($M)", values=ebitda,
                       is_calculated=True, is_subtotal=True,
                       note="= Operating Income + D&A")

    # EBITDA margin
    ebitda_margin = {}
    for year in YEARS_ACTUAL:
        e = ebitda.get(year)
        r = HISTORICAL["revenue"].get(year)
        if e and r:
            ebitda_margin[year] = e / r

    row = add_data_row(row, "  EBITDA Margin (%)", values=ebitda_margin, fmt="pct",
                       is_calculated=True)

    # ROE
    roe = {}
    for year in YEARS_ACTUAL:
        ni = HISTORICAL["net_income"].get(year)
        eq = HISTORICAL["total_equity"].get(year)
        if ni and eq and eq != 0:
            roe[year] = ni / eq

    row = add_data_row(row, "  ROE (%)", values=roe, fmt="pct",
                       is_calculated=True)

    # ROIC approximation
    roic = {}
    for year in YEARS_ACTUAL:
        oi = HISTORICAL["operating_income"].get(year)
        r = HISTORICAL["revenue"].get(year)
        ta = HISTORICAL["total_assets"].get(year)
        cash_val = HISTORICAL["cash"].get(year)
        if oi and ta and cash_val:
            invested = ta - cash_val
            if invested != 0:
                nopat = oi * 0.825  # (1 - 17.5% tax)
                roic[year] = nopat / invested

    row = add_data_row(row, "  ROIC (%)", values=roic, fmt="pct",
                       is_calculated=True,
                       note="= NOPAT / Invested Capital")

    # =========================================================================
    # FREEZE PANES & PRINT SETTINGS
    # =========================================================================
    ws.freeze_panes = "C5"  # Freeze labels + year headers

    return ws


# =============================================================================
# MAIN
# =============================================================================
def main():
    print("Building PayPal Financial Model (Cover + Assumptions)...\n")

    wb = Workbook()
    build_cover(wb)
    build_assumptions(wb)

    # Create placeholder sheets for navigation
    for name, color in [
        ("Income Statement", "28A745"),
        ("Balance Sheet", "17A2B8"),
        ("Cash Flow", "FFC107"),
        ("Ratios", "6F42C1"),
        ("DCF", "DC3545"),
        ("Scenarios", "FF6B35"),
        ("Investment Memo", "2E4057"),
    ]:
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = color
        ws.cell(row=2, column=2, value=f"{name} — To be built in next step").font = Font(
            name="Arial", size=14, italic=True, color="999999"
        )

    wb.save(OUTPUT_PATH)
    print(f"  ✓ Saved: {OUTPUT_PATH}")
    print(f"  ✓ Tabs: {', '.join(wb.sheetnames)}")
    print(f"\n  Open in Excel and review the Assumptions tab.")
    print(f"  Blue values = editable inputs. Yellow cells = key forecast assumptions.")
    print(f"  Adjust any forecast assumptions you disagree with before we build the statements.\n")


if __name__ == "__main__":
    main()
